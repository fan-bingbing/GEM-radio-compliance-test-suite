import visa
from openpyxl import Workbook
import time
import sys
#from openpyxl.drawing.image import Image

rm = visa.ResourceManager()
SMB = rm.open_resource('USB0::0x0AAD::0x0054::106409::INSTR') # SMB100A
SIGEN_file_read = Workbook() # open a excel file
ws = SIGEN_file_read.active # assign active sheet to ws
ws.title = "FREQ" # rename ws

#img = Image("ACP01.bmp") # this requres install pillow to be able to fetch image

ws.cell(row = 1, column = 1, value = "Frequency") # give tile for each column in first row in ws
ws.cell(row = 1, column = 2, value = "Level")
ws.cell(row = 1, column = 3, value = "Unit")
ws.cell(row = 1, column = 4, value = "ACP Plot")

#ws.add_image(img, "D1")

for i in range(0, 5):# try reading frequecny/level from SMB100A 5 times with 1s delay for each time

    SIGEN_freq = float((SMB.query("SOUR:FREQ?"))) # transfer string to float
    SIGEN_level = float((SMB.query("SOUR:POW?"))) + 107 # transfer dBm to dBuV

    ws.cell(row = i+2, column = 1, value = SIGEN_freq) # write values in 3 columns per row iteration
    ws.cell(row = i+2, column = 2, value = SIGEN_level)
    ws.cell(row = i+2, column = 3, value = "dBuV")

    sys.stdout.write('\r')# add progress bar
    sys.stdout.write("[%-5s] %d%%" % ('='*(i+1), 20*(i+1)))
    sys.stdout.flush()
    print(f"-----{SIGEN_freq / 1e6} MHz, {SIGEN_level} dBuV") # print values got from each step

    time.sleep(1) # 1 second dealy on each step

SIGEN_file.save("SIGEN_Read.xlsx") # save excel file
# just for first commit
