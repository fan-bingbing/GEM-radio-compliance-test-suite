import visa
import time
from openpyxl import load_workbook

# naming rules for excel:
# 1. for operating variables related to Test_Setup.xlsx: use SFile_write, SSheet
# 2. for operating variables related to Test_Result.xlsx: use RFile_write, RSheet

def FreqError_Power(Test_frequency):
    rm = visa.ResourceManager()
    FSV = rm.open_resource('TCPIP0::192.168.10.9::hislip0::INSTR') # Spec An
    # below codes are for setting test frequency in Test_Setup.xlsx according to user's input
    SFile_write = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file
    SSheet = SFile_write["Freq_Error_Power"] # load existing SSheet named "ACP"
    SSheet.cell(row = 1, column = 2, value = Test_frequency) # write test frequency in this SSheet
    SFile_write.save("Test_Setup.xlsx") # save existing .xlsx file
    # above codes are for setting test frequency in Test_Setup.xlsx according to user's input

    RFile_write = load_workbook(filename = "Test_Result.xlsx") # load Test_Result.xlsx
    RSheet = RFile_write["Ferror_Pow"] # load "Ferror_Pow" sheet in .xlsx


    Centre_frequency = SSheet["B1"].value #
    Span_frequency = SSheet["B2"].value #
    RBW = SSheet["B3"].value #
    VBW = SSheet["B4"].value #
    RF_level = SSheet["B5"].value #
    Attenuation = SSheet["B6"].value #
    RefLev_offset = SSheet["B7"].value#
    Trace_peak = SSheet["B8"].value #
    Transducer1 = SSheet["B9"].value # cable
    Trans1_ON = SSheet["B10"].value
    Transducer2 = SSheet["B11"].value # 30dB attenuator
    Trans2_ON = SSheet["B12"].value
    Transducer3 = SSheet["B13"].value # High pass filter
    Trans3_ON = SSheet["B14"].value
    Limit_line_1 = SSheet["B15"].value # ASNZS4365:2011
    Limit_line_1_ON = SSheet["B16"].value # ASNZS4365:2011
    Limit_line_2 = SSheet["B17"].value # ASNZS4295:2015
    Limit_line_2_ON = SSheet["B18"].value # ASNZS4295:2015
    Sweep_points= SSheet["B19"].value # get sweep points

    FSV.write(f"*RST")
    FSV.write("SYST:DISP:UPD ON")
    FSV.write(f"FREQ:CENT {Centre_frequency}MHz")
    FSV.write(f"FREQ:SPAN {Span_frequency}Hz")
    FSV.write(f"BAND {RBW}Hz")
    FSV.write(f"BAND:VID {VBW}Hz")
    FSV.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
    FSV.write(f"DISP:TRAC:Y:RLEV {RF_level}")
    FSV.write(f"INP:ATT {Attenuation}")
    FSV.write(f"{RefLev_offset}")
    FSV.write(f"{Trace_peak}")
    FSV.write(f"CORR:TRAN:SEL '{Transducer1}'")
    FSV.write(f"CORR:TRAN {Trans1_ON} ")
    FSV.write(f"CORR:TRAN:SEL '{Transducer2}'")
    FSV.write(f"CORR:TRAN {Trans2_ON}")
    FSV.write(f"CORR:TRAN:SEL '{Transducer3}'")
    FSV.write(f"CORR:TRAN {Trans3_ON}")
    FSV.write(f"CALC:LIM:NAME '{Limit_line_1}'")
    FSV.write(f"CALC:LIM:UPP:STAT {Limit_line_1_ON}")
    FSV.write(f"CALC:LIM:NAME '{Limit_line_2}'")
    FSV.write(f"CALC:LIM:UPP:STAT {Limit_line_2_ON}")
    FSV.write(f"SWE:POIN {Sweep_points}")
    FSV.write(f"DISP:TRAC:MODE MAXH")
    time.sleep(5)# wait 5 seconds for trace to stable
    FSV.write(f"DISP:TRAC:MODE VIEW")
    FSV.query("*OPC?")

    FSV.write("CALC:MARK1:MAX")
    Frequency = FSV.query("CALC:MARK1:X?")
    Level = float(FSV.query("CALC:MARK1:Y?"))
    Frequency_error = float(Frequency) - float(Centre_frequency)*1e6
    indication = (FSV.query("*OPC?")).replace("1","Frequency error test Completed") # replace return character "1" to "completed"
    RSheet.cell(row = 2, column = 1, value = Frequency_error)
    RSheet.cell(row = 2, column = 2, value = Level)
    RFile_write.save("Test_Result.xlsx") # save existing .xlsx file
    FSV.close()

    return {'Frequency_error':Frequency_error, 'Carrier_power':Level, 'Indication':indication}
    # return multiple values using dictionary
