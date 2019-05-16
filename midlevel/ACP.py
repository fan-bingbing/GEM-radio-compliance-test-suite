import visa
from openpyxl import load_workbook

import time
import re




def Adjacent_channel_power(Test_frequency):

    rm = visa.ResourceManager()
    FSV = rm.open_resource('TCPIP0::192.168.10.9::hislip0::INSTR') # Spec An
    SML = rm.open_resource('ASRL4::INSTR') # wanted Signal

# below codes are for setting test frequency in Test_Setup.xlsx according to user's input
    FSV_file_write = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file
    sheet = FSV_file_write["ACP"] # load existing sheet named "ACP"
    sheet.cell(row = 1, column = 2, value = Test_frequency) # write test frequency in this sheet
    FSV_file_write.save("Test_Setup.xlsx") # save existing .xlsx file
# above codes are for setting test frequency in Test_Setup.xlsx according to user's input

    SML.clear()  # Clear instrument io buffers and status
    FSV.clear()

    FSV_file_write = load_workbook(filename = "Test_Setup.xlsx") # create a workbook from existing .xlsx file
    sheet = FSV_file_write["ACP"] # load setup sheet in .xlsx to sheet

    # following code is to initialize SML
    Frequency_AF = sheet["G1"].value #
    Level_AF = sheet["G2"].value #
    AF_output_on = sheet["G3"].value #
    SML.write(f"*RST")
    SML.write("SYST:DISP:UPD ON")
    SML.write(f":FM:INT:FREQ {Frequency_AF}kHz")
    SML.write(f":OUTP2:VOLT {Level_AF}mV")
    SML.write(f":OUTP2 {AF_output_on}")
    SML.query('*OPC?')
    time.sleep(1)
    # above code is to initialize SML

    # following code is to initialize FSV
    Centre_frequency = sheet["K1"].value #
    Dev_PerDivision = sheet["K2"].value #
    Demod_BW = sheet["K3"].value #
    AF_Couple = sheet["K4"].value #
    RF_level = sheet["K5"].value #
    Attenuation = sheet["K6"].value # get attenuation
    RefLev_offset = sheet["K7"].value# get RFlevel offset
    Trace_Peak = sheet["K8"].value # get trace to Pos peak
    Demod_MT = sheet["K9"].value #
    Cont_sweep = sheet["K10"].value #
    FSV.write(f"*RST")
    FSV.write("SYST:DISP:UPD ON")
    FSV.write("ADEM ON")
    FSV.write(f"FREQ:CENT {Centre_frequency}MHz")
    FSV.write(f"DISP:TRAC:Y:PDIV {Dev_PerDivision}kHz")
    FSV.write(f"BAND:DEM {Demod_BW}kHz")
    FSV.write(f"ADEM:AF:COUP {AF_Couple}")#Set AF coupling to AC, the frequency offset is automatically corrected.
    # i.e. the trace is always symmetric with respect to the zero line
    FSV.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
    FSV.write(f"DISP:TRAC:Y:RLEV {RF_level}")
    FSV.write(f"INP:ATT {Attenuation}")
    FSV.write(f"{Trace_Peak}")
    FSV.write(f"ADEM:MTIM {Demod_MT}ms")
    FSV.write(f"INIT:CONT {Cont_sweep}")
    FSV.query("*OPC?")
    time.sleep(1)
    # above code is to initialize FSV

    # following code is to find Audio output level satisfying standard condtion (around 1.5kHz deviation)
    FSV.write(f"INIT:CONT OFF")
    FSV.query('*OPC?')
    time.sleep(1)
    Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0 #get the initial deviation value

    if Dev_Reading < 1.5:
        while Dev_Reading < 1.42:
            Level_AF = Level_AF+1
            SML.write(f":OUTP2:VOLT {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0
    else:
        while Dev_Reading > 1.58:
            Level_AF = Level_AF-1
            SML.write(f":OUTP2:VOLT {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0
    # above code is to find Audio output level satisfying standard condtion (around 1.5kHz deviation)

    SML.query('*OPC?')

    # following code is to initialize FSV to complete the test
    Centre_frequency = sheet["B1"].value # get all parameters from Test_Setup.xlsx
    Span_frequency = sheet["B2"].value #
    RBW = sheet["B3"].value #
    VBW = sheet["B4"].value #
    RF_level = sheet["B5"].value #
    Attenuation = sheet["B6"].value #
    RefLev_offset = sheet["B7"].value#
    Trace_RMS = sheet["B8"].value #

    Tx_CHBW = sheet["B10"].value
    AJ_CHBW = sheet["B11"].value
    AT_CHBW = sheet["B12"].value
    AJ_CHNUM = sheet["B13"].value
    AJ_SPACE = sheet["B14"].value
    AT_SPACE = sheet["B15"].value
    Power_Mode = sheet["B16"].value
    Ave_number = sheet["B17"].value

    FSV.write(f"*RST")
    FSV.write("SYST:DISP:UPD ON")
    FSV.write("CALC:MARK:FUNC:POW:SEL ACP")

    FSV.write(f"FREQ:CENT {Centre_frequency}MHz") # set all parameters
    FSV.write(f"FREQ:SPAN {Span_frequency}kHz")
    FSV.write(f"BAND {RBW}Hz")
    FSV.write(f"BAND:VID {VBW}Hz")
    FSV.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
    FSV.write(f"DISP:TRAC:Y:RLEV {RF_level}")
    FSV.write(f"INP:ATT {Attenuation}")
    FSV.write(f"{Trace_RMS}")

    FSV.write(f"POW:ACH:BWID:CHAN1 {Tx_CHBW}kHz")
    FSV.write(f"POW:ACH:BWID:ACH {AJ_CHBW}kHz")
    FSV.write(f"POW:ACH:BWID:ALT1 {AT_CHBW}kHz")
    FSV.write(f"POW:ACH:ACP {AJ_CHNUM}")
    FSV.write(f"POW:ACH:SPAC {AJ_SPACE}kHz")
    FSV.write(f"POW:ACH:SPAC:ALT1 {AT_SPACE}kHz")
    FSV.write(f"POW:ACH:MODE {Power_Mode}")
    FSV.write(f"SWE:COUN {Ave_number}")
    FSV.write(f"CALC:MARK:FUNC:POW:MODE WRIT")
    FSV.write(f"DISP:TRAC:MODE AVER")
    time.sleep(10)
    FSV.write(f"DISP:TRAC:MODE VIEW")

    print(FSV.query("CALC:MARK:FUNC:POW:RES? ACP"))# query ACP data result
    #re.sub("\D", "", s) question: how to strip out 4 numbers from a string and put them into a list?


    indication = (FSV.query("*OPC?")).replace("1","ACP test Completed") # replace return character "1" to "completed"
    SML.write(":OUTP2 OFF")# turn off audio output at the end of the test
    FSV.close()
    SML.close()
    return indication
