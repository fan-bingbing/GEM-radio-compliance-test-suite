import visa
import time
from openpyxl import load_workbook

# naming rules for excel:
# 1. for operating variables related to Test_Setup.xlsx: use SFile_write, SSheet
# 2. for operating variables related to Test_Result.xlsx: use RFile_write, RSheet


def Max_Deviation(Test_frequency):
    rm = visa.ResourceManager()
    FSV = rm.open_resource('TCPIP0::192.168.10.9::hislip0::INSTR') # Spec An
    SML = rm.open_resource('ASRL4::INSTR') # wanted Signal
    AF_list = [100, 300, 500, 700, 900, 1000, 1300, 1500, 1700, 1900, 2000, 2300, 2550] # audio frequency list
    Reading_list = [] # empty list for deviation result storage

    # below codes are for setting test frequency in Test_Setup.xlsx according to user's input
    SFile_write = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file
    SSheet = SFile_write["Max_Deviation"] # load existing SSheet named "ACP"
    SSheet.cell(row = 1, column = 2, value = Test_frequency) # write test frequency in this SSheet
    SFile_write.save("Test_Setup.xlsx") # save existing .xlsx file
    # above codes are for setting test frequency in Test_Setup.xlsx according to user's input


    SML.clear()  # Clear instrument io buffers and status
    FSV.clear()

    SFile_write = load_workbook(filename = "Test_Setup.xlsx") # create a workbook from existing .xlsx file
    SSheet = SFile_write["Max_Deviation"] # load setup SSheet in .xlsx to SSheet

    # following code is to initialize SML
    Frequency_AF = SSheet["G1"].value #
    Level_AF = SSheet["G2"].value #
    AF_output_on = SSheet["G3"].value #
    SML.write(f"*RST")
    SML.write("SYST:DISP:UPD ON")
    SML.write(f":FM:INT:FREQ {Frequency_AF}kHz")
    SML.write(f":OUTP2:VOLT {Level_AF}mV")
    SML.write(f":OUTP2 {AF_output_on}")
    SML.query('*OPC?')
    time.sleep(1)
    # above code is to initialize SML

    # following code is to initialize FSV
    Centre_frequency = SSheet["B1"].value #
    Dev_PerDivision = SSheet["B2"].value #
    Demod_BW = SSheet["B3"].value #
    AF_Couple = SSheet["B4"].value #
    RF_level = SSheet["B5"].value #
    Attenuation = SSheet["B6"].value # get attenuation
    RefLev_offset = SSheet["B7"].value# get RFlevel offset
    Trace_Peak = SSheet["B8"].value # get trace to Pos peak
    Demod_MT = SSheet["B9"].value #
    Cont_sweep = SSheet["B10"].value #
    FSV.write(f"*RST")
    FSV.write("SYST:DISP:UPD ON")
    FSV.write("ADEM ON")
    FSV.write(f"FREQ:CENT {Centre_frequency}MHz")
    FSV.write(f"DISP:TRAC:Y:PDIV {Dev_PerDivision}kHz")
    FSV.write(f"BAND:DEM {Demod_BW}kHz")
    FSV.write(f"ADEM:AF:COUP {AF_Couple}")#Set AF coupling to AC, the frequency offset is automatically corrected.
    # i.e. the trace is always symmetric with respect to the zero line
    FSV.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
    FSV.write(f"DISP:TRAC:Y:RLEV {RF_level}")
    FSV.write(f"INP:ATT {Attenuation}")
    FSV.write(f"{Trace_Peak}")
    FSV.write(f"ADEM:MTIM {Demod_MT}ms")
    FSV.write(f"INIT:CONT {Cont_sweep}")
    FSV.query("*OPC?")
    time.sleep(1)
    # above code is to initialize FSV

    # following code is to find Audio output level satisfying standard condtion (around 1.5kHz deviation)
    FSV.write(f"INIT:CONT OFF")
    FSV.query('*OPC?')
    time.sleep(1)
    Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0 #get the initial deviation value

    if Dev_Reading < 1.5:
        while Dev_Reading < 1.42:
            Level_AF = Level_AF+1
            SML.write(f":OUTP2:VOLT {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0
    else:
        while Dev_Reading > 1.58:
            Level_AF = Level_AF-1
            SML.write(f":OUTP2:VOLT {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0
    # above code is to find Audio output level satisfying standard condtion (around 1.5kHz deviation)

    SML.query('*OPC?')

    RFile_write = load_workbook(filename = "Test_Result.xlsx") # load Test_Result.xlsx
    RSheet = RFile_write["Max_Dev"] # load "Max_Dev" sheet in .xlsx

    #following code is to vary audio frequency to complete the test
    Level_AF = 100*Level_AF # bring audio level up 20dB in one step according to standard
    SML.write(f":OUTP2:VOLT {Level_AF}mV")
    for i in range(0,13):
        print(f"At Audio frequency:{AF_list[i]}")
        SML.write(f":FM:INT:FREQ {AF_list[i]}Hz")
        SML.query('*OPC?')
        FSV.write(f"INIT:CONT OFF")
        time.sleep(1)
        Dev = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000
        Reading_list.append(Dev)
        FSV.query("*OPC?")
        FSV.write(f"INIT:CONT ON")
        print(f"Deviation is {Reading_list[i]}kHz")
        RSheet.cell(row = i+2, column = 1, value = AF_list[i])
        RSheet.cell(row = i+2, column = 2, value = Dev)

    RFile_write.save("Test_Result.xlsx") # save existing .xlsx file
    SML.write(":OUTP2 OFF")# turn off audio output at the end of the test
    indication = (FSV.query("*OPC?")).replace("1","Completed")
    Reading_list.append(indication)
    FSV.close()
    SML.close()
    return Reading_list # results are stored in a list
    #above code is to vary audio frequency to complete the test
