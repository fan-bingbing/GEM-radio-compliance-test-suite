import visa
from openpyxl import load_workbook
import time
import re

# naming rules for excel:
# 1. for operating variables related to Test_Setup.xlsx: use SFile_write, SSheet
# 2. for operating variables related to Test_Result.xlsx: use RFile_write, RSheet

def Blocking_immunity(Test_frequency):

    rm = visa.ResourceManager()
    SML = rm.open_resource('ASRL4::INSTR') # wanted Signal
    SMB = rm.open_resource('USB0::0x0AAD::0x0054::106409::INSTR') # Unwanted Signal
    #FSV = rm.open_resource('TCPIP0::192.168.10.9::hislip0::INSTR') # Spec An
    CMS = rm.open_resource('GPIB0::24::INSTR') # Audio Analyzer

    # below codes are for setting test frequency in Test_Setup.xlsx according to user's input
    SFile_write = load_workbook(filename = "Test_Setup.xlsx") # create a workbook from existing .xlsx file
    SSheet = SFile_write["Blocking"] # load setup sheet in .xlsx to sheet
    SSheet.cell(row = 1, column = 3, value = Test_frequency) # write test frequency in this sheet
    SFile_write.save("Test_Setup.xlsx") # save existing .xlsx file
    # above codes are for setting test frequency in Test_Setup.xlsx according to user's input

    SML.clear()  # Clear instrument io buffers and status
    SMB.clear()
    CMS.clear()
    # below codes are for setting standard test condition SML(FM) = 15.5dBuV
    SFile_write = load_workbook(filename = "Test_Setup.xlsx") # create a workbook from existing .xlsx file
    SSheet = SFile_write["Blocking"] # load setup sheet in .xlsx to sheet

    Frequency_RF = SSheet["C1"].value #
    Level_RF = SSheet["C2"].value #
    Frequency_AF = SSheet["C3"].value #
    Deviation = SSheet["C4"].value #
    Mod_state = SSheet["C5"].value #
    RF_power_on = SSheet["C6"].value #

    SML.write(f"*RST")
    SML.write("SYST:DISP:UPD ON")
    SML.write(f"FREQ {Frequency_RF}MHz")
    SML.write(f":POW:UNIT dBuV")
    SML.write(f":POW {Level_RF}dBuV")
    SML.write(f":FM:INT:FREQ {Frequency_AF}kHz")
    SML.write(f":FM:DEV {Deviation}kHz")
    SML.write(f":FM:STAT {Mod_state}")
    SML.write(f":OUTP1 {RF_power_on}")
    # above codes are for setting standard test condition SML(FM) = 15.5dBuV
    SML.query('*OPC?')
    # below code block are for setting start point of SMB(interference) = 50dBuV
    Frequency_RF1 = SSheet["C8"].value #
    Frequency_RF2 = SSheet["E9"].value #
    Level_RF = SSheet["C9"].value #
    Frequency_AF = SSheet["C10"].value #
    Deviation = SSheet["C11"].value #
    Mod_state = SSheet["C12"].value #
    RF_power_on = SSheet["C13"].value #

    SMB.write(f"*RST")
    #SMB.write("SYST:DISP:UPD ON")
    SMB.write(f":FREQ {Frequency_RF1}MHz")

    SMB.write(f":UNIT:POW dBuV")
    SMB.write(f":POW {Level_RF}")
    SMB.write(f":FM:INT:FREQ {Frequency_AF}Hz")
    SMB.write(f":FM:DEV {Deviation}kHz")
    SMB.write(f":FM:STAT {Mod_state}")
    SMB.write(f":OUTP1 {RF_power_on}")
    # above code block are for setting start point of SMB(interference) = 50dBuV
    SMB.query('*OPC?')

    RFile_write = load_workbook(filename = "Test_Result.xlsx") # load Test_Result.xlsx
    RSheet = RFile_write["Blocking"] # load "ACP" sheet in .xlsx

    SINAD_data_str = CMS.query("SINAD:R?") # get initial SINAD value
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(f"Initial SINAD value for BLK+:{SINAD_data_num}")

    # below code block are for ACS high side test
    i=1
    while float(SINAD_data_num) > 14.0:
        RSheet.cell(row = i+1, column = 1, value = Level_RF)
        RSheet.cell(row = i+1, column = 2, value = SINAD_data_num)
        Level_RF = Level_RF + 1
        SMB.write(f":POW {Level_RF}")
        SMB.query('*OPC?')
        SINAD_data_str = CMS.query("SINAD:R?")
        SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
        print(f"SINAD+:{SINAD_data_num}")
        i = i+1
    RSheet.cell(row = i+1, column = 1, value = Level_RF)# complete last row entry
    RSheet.cell(row = i+1, column = 2, value = SINAD_data_num)# complete last row entry
    BLK_high = float(SMB.query(f":POW? "))+3.5-(15.5-3.5)
    RSheet.cell(row = 2, column = 3, value = BLK_high)
    RFile_write.save("Test_Result.xlsx") # save existing .xlsx file
    # above code block are for ACS high side test

    SMB.query('*OPC?')
    time.sleep(5)
    SMB.write(f":FREQ {Frequency_RF2}MHz")# prepare test again on the other side
    Level_RF = SSheet["C9"].value # reset SMB level to original setting
    SMB.write(f":POW {Level_RF}")
    SINAD_data_str = CMS.query("SINAD:R?") # get initial SINA value
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(f"Initial SINAD value for BLK-:{SINAD_data_num}")

    # below code block are for ACS low side test
    i=1
    while float(SINAD_data_num) > 14.0:
        RSheet.cell(row = i+1, column = 4, value = Level_RF)
        RSheet.cell(row = i+1, column = 5, value = SINAD_data_num)
        Level_RF = Level_RF + 1
        SMB.write(f":POW {Level_RF}")
        SMB.query('*OPC?')
        SINAD_data_str = CMS.query("SINAD:R?")
        SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
        print(f"SINAD-:{SINAD_data_num}")
        i = i+1
    RSheet.cell(row = i+1, column = 4, value = Level_RF)# complete last row entry
    RSheet.cell(row = i+1, column = 5, value = SINAD_data_num)# complete last row entry
    BLK_low = float(SMB.query(f":POW? "))+3.5-(15.5-3.5)
    RSheet.cell(row = 2, column = 6, value = BLK_low)
    RFile_write.save("Test_Result.xlsx") # save existing .xlsx file
    # above code block are for ACS low side test

    SMB.query('*OPC?')
    indication = (SMB.query("*OPC?")).replace("1","Blocking test Completed")

    SML.close()
    SMB.close()
    #FSV.clear()
    CMS.close()

    return {'BLK_high':BLK_high, 'BLK_low':BLK_low, 'Indication':indication}
