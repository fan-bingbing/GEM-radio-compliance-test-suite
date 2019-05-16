import visa
import time
from openpyxl import load_workbook
import config # contains global variables

def Conducted_spurious(setup_sheet):
    rm = visa.ResourceManager()
    FSV = rm.open_resource('TCPIP0::192.168.10.9::hislip0::INSTR') # Spec An

    # below codes are for selecting correct limit line based on global varable limit_line_factor
    if config.limit_line_factor == 2:
        FSV_file_write = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file
        sheet = FSV_file_write[setup_sheet] # load existing sheet named "ACP"
        sheet.cell(row = 16, column = 2, value = "ON") # write test frequency in this sheet
        sheet.cell(row = 18, column = 2, value = "OFF") # write test frequency in this sheet
        FSV_file_write.save("Test_Setup.xlsx") # save existing .xlsx file
    else:
        FSV_file_write = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file
        sheet = FSV_file_write[setup_sheet] # load existing sheet named "ACP"
        sheet.cell(row = 16, column = 2, value = "OFF") # write test frequency in this sheet
        sheet.cell(row = 18, column = 2, value = "ON") # write test frequency in this sheet
        FSV_file_write.save("Test_Setup.xlsx") # save existing .xlsx file

    # above codes are for selecting correct limit line based on global varable limit_line_factor

    FSV.clear()

    FSV_file_write = load_workbook(filename = "Test_Setup.xlsx") # create a workbook from existing .xlsx file
    sheet = FSV_file_write[setup_sheet] # load setup sheet in .xlsx to sheet

    start_frequency = sheet["B1"].value # get start frequency value from sheet
    stop_frequency = sheet["B2"].value # get stop frequency value from sheet
    RBW = sheet["B3"].value # get RBW value from sheet
    VBW = sheet["B4"].value # get VBW value from sheet
    RF_level = sheet["B5"].value # get RF_level
    Attenuation = sheet["B6"].value # get attenuation
    RefLev_offset = sheet["B7"].value# get RFlevel offset
    Trace_peak = sheet["B8"].value # get trace to Pos peak
    Transducer1 = sheet["B9"].value # cable
    Trans1_ON = sheet["B10"].value
    Transducer2 = sheet["B11"].value # 30dB attenuator
    Trans2_ON = sheet["B12"].value
    Transducer3 = sheet["B13"].value # High pass filter
    Trans3_ON = sheet["B14"].value
    Limit_line_1 = sheet["B15"].value # ASNZS4365:2011
    Limit_line_1_ON = sheet["B16"].value # ASNZS4365:2011
    Limit_line_2 = sheet["B17"].value # ASNZS4365:2011
    Limit_line_2_ON = sheet["B18"].value # ASNZS4365:2011
    Sweep_points= sheet["B19"].value # get sweep points


    FSV.write(f"*RST")
    FSV.write("SYST:DISP:UPD ON")
    FSV.write(f"FREQ:STAR {start_frequency}MHz")
    FSV.write(f"FREQ:STOP {stop_frequency}MHz")
    FSV.write(f"BAND {RBW}kHz")
    FSV.write(f"BAND:VID {VBW}kHz")
    FSV.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
    FSV.write(f"DISP:TRAC:Y:RLEV {RF_level}")
    FSV.write(f"INP:ATT {Attenuation}")
    FSV.write(f"{RefLev_offset}")
    FSV.write(f"{Trace_peak}")
    FSV.write(f"CORR:TRAN:SEL '{Transducer1}'")
    FSV.write(f"CORR:TRAN {Trans1_ON} ")
    FSV.write(f"CORR:TRAN:SEL '{Transducer2}'")
    FSV.write(f"CORR:TRAN {Trans2_ON}")
    FSV.write(f"CORR:TRAN:SEL '{Transducer3}'")
    FSV.write(f"CORR:TRAN {Trans3_ON}")
    FSV.write(f"CALC:LIM:NAME '{Limit_line_1}'")
    FSV.write(f"CALC:LIM:UPP:STAT {Limit_line_1_ON}")
    FSV.write(f"CALC:LIM:NAME '{Limit_line_2}'")
    FSV.write(f"CALC:LIM:UPP:STAT {Limit_line_2_ON}")
    FSV.write(f"SWE:POIN {Sweep_points}")
    FSV.write(f"DISP:TRAC:MODE MAXH")
    time.sleep(5)# wait 10 seconds for trace to stable
    FSV.write(f"DISP:TRAC:MODE VIEW")
    FSV.query("*OPC?")

    FSV.write("CALC:MARK1:MAX")
    Frequency = float(FSV.query("CALC:MARK1:X?"))
    Level = float(FSV.query("CALC:MARK1:Y?"))

    indication = (FSV.query("*OPC?")).replace("1","Completed") # replace return character "1" to "completed"
    FSV.close()

    return {'Mark frequency':Frequency, 'Mark level':Level, 'Indication':indication}
