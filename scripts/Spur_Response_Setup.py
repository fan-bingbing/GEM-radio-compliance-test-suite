import visa
from openpyxl import load_workbook
import re
import time

# naming rules for excel:
# 1. for operating variables related to Test_Setup.xlsx: use SFile_write, SSheet
# 2. for operating variables related to Test_Result.xlsx: use RFile_write, RSheet



rm = visa.ResourceManager()
SML = rm.open_resource('ASRL4::INSTR') # wanted Signal
SMB = rm.open_resource('USB0::0x0AAD::0x0054::106409::INSTR') # Unwanted Signal
CMS = rm.open_resource('GPIB0::24::INSTR') # Audio Analyzer

#scope.write_termination = '\n'
SML.clear()  # Clear instrument io buffers and status
SMB.clear()
CMS.clear()

SFile_write = load_workbook(filename = "Test_Setup.xlsx") # open "Test_Setup.xlsx"
SSheet = SFile_write["Spurious_Response"] # load "Spurious_Response" sheet

# below code block to configure SML to standard teset condition
Frequency_RF = SSheet["C1"].value #
Level_RF = SSheet["C2"].value #
Frequency_AF = SSheet["C3"].value #
Deviation = SSheet["C4"].value #
Mod_state = SSheet["C5"].value #
RF_power_on = SSheet["C6"].value #

SML.write(f"*RST")
SML.write("SYST:DISP:UPD ON")
SML.write(f"FREQ {Frequency_RF}MHz")
SML.write(f":POW:UNIT dBuV")
SML.write(f":POW {Level_RF}dBuV")
SML.write(f":FM:INT:FREQ {Frequency_AF}kHz")
SML.write(f":FM:DEV {Deviation}kHz")
SML.write(f":FM:STAT {Mod_state}")
SML.write(f":OUTP1 {RF_power_on}")

SML.query('*OPC?')
# above code block to configure SML to standard teset condition

# below code block to configure SMB to initial status
Frequency_RF = SSheet["C8"].value #
Level_RF = SSheet["C9"].value #
Frequency_AF = SSheet["C10"].value #
Deviation = SSheet["C11"].value #
Mod_state = SSheet["C12"].value #
RF_power_on = SSheet["C13"].value #

SMB.write(f"*RST")
#SMB.write("SYST:DISP:UPD ON")
SMB.write(f":FREQ {Frequency_RF}MHz")
SMB.write(f":UNIT:POW dBuV")
SMB.write(f":POW {Level_RF}")
SMB.write(f":FM:INT:FREQ {Frequency_AF}Hz")
SMB.write(f":FM:DEV {Deviation}kHz")
SMB.write(f":FM:STAT {Mod_state}")
SMB.write(f":OUTP1 {RF_power_on}")

SMB.query('*OPC?')
# above code block to configure SMB to initial status

RFile_write = load_workbook(filename = "Test_Result.xlsx") # load Test_Result.xlsx
RSheet = RFile_write["Spurious_Response"] # load "Spurious_Response" sheet

# read initial SINAD
SINAD_data_str = CMS.query("SINAD:R?")
#print(SINAD_data_str)
SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
print(SINAD_data_num)

# below code block to test Spurious_Response
for i in range(0,100):
    if float(SINAD_data_num) > 14.0:
        RSheet.cell(row = i+2, column = 1, value = Level_RF)
        RSheet.cell(row = i+2, column = 2, value = SINAD_data_num)
        Level_RF = Level_RF + 1
        SMB.write(f":POW {Level_RF}")
        SMB.query('*OPC?')
        SINAD_data_str = CMS.query("SINAD:R?")
        SINAD_data_num = re.findall(r'\d', SINAD_data_str)[0]# handle return value of 0
        if SINAD_data_num != '0':
            SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
            print(SINAD_data_num)
        else:
            break
    else:
        break

Spurious_Response = float(SMB.query(f":POW? "))-(45.5-6) # wanted signal expressed as dBuV emf
RSheet.cell(row = 2, column = 3, value = Spurious_Response)
print(f"Spurious_Response+:{Spurious_Response}")
# above code block to test Spurious_Response


RFile_write.save("Test_Result.xlsx")# save result
SML.close()
SMB.close()
CMS.close()
