import visa
from openpyxl import load_workbook
import re
import time
import math

# naming rules for excel:
# 1. for operating variables related to Test_Setup.xlsx: use SFile_write, SSheet
# 2. for operating variables related to Test_Result.xlsx: use RFile_write, RSheet



rm = visa.ResourceManager()
SML = rm.open_resource('ASRL4::INSTR') # wanted Signal
#SMB = rm.open_resource('USB0::0x0AAD::0x0054::106409::INSTR') # Unwanted Signal
CMS = rm.open_resource('GPIB0::24::INSTR') # Audio Analyzer

#scope.write_termination = '\n'
SML.clear()  # Clear instrument io buffers and status
CMS.clear()

SFile_write = load_workbook(filename = "Test_Setup.xlsx") # open "Test_Setup.xlsx"
SSheet = SFile_write["Noise_Hum"] # load "Noise_Hum" sheet

# below code block to configure SML to standard teset condition
Frequency_RF = SSheet["C1"].value #
Level_RF = SSheet["C2"].value #
Frequency_AF = SSheet["C3"].value #
Deviation = SSheet["C4"].value #
Mod_state = SSheet["C5"].value #
RF_power_on = SSheet["C6"].value #

SML.write(f"*RST")
SML.write("SYST:DISP:UPD ON")
SML.write(f"FREQ {Frequency_RF}MHz")
SML.write(f":POW:UNIT dBuV")
SML.write(f":POW {Level_RF}dBuV")
SML.write(f":FM:INT:FREQ {Frequency_AF}kHz")
SML.write(f":FM:DEV {Deviation}kHz")
SML.write(f":FM:STAT {Mod_state}")
SML.write(f":OUTP1 {RF_power_on}")

SML.query('*OPC?')
# above code block to configure SML to standard teset condition

RFile_write = load_workbook(filename = "Test_Result.xlsx") # load Test_Result.xlsx
RSheet = RFile_write["Noise_Hum"] # load "Noise_Hum" sheet

# below code block to test audio output level with modulation
SINAD_data_str = CMS.query("LE:A:R?")# get audio level output
print(SINAD_data_str)
RSheet.cell(row = 2, column = 2, value = SINAD_data_str)
SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]#strip out number
print(f"{SINAD_data_num}mV")
AFLevel_Mod = float(SINAD_data_num) # AF level in mV result from with modulation
RSheet.cell(row = 2, column = 3, value = AFLevel_Mod)
# above code block to test audio output level with modulation



# below code block to test audio output level without modulation
SML.write(f":FM:STAT OFF")# switch off modulation
time.sleep(3) # wait 3 seconds for CMS reading to be stable
SINAD_data_str = CMS.query("LE:A:R?")# get audio level output
print(SINAD_data_str)
RSheet.cell(row = 3, column = 2, value = SINAD_data_str)
SINAD_data_num = re.findall(r'\d+', SINAD_data_str)[0]
print(f"{SINAD_data_num}uV")
AFLevel_Unmod = float(SINAD_data_num) # AF level in uV result from without modulation
RSheet.cell(row = 3, column = 3, value = AFLevel_Unmod/1000)
# above code block to test audio output level without modulation



result = 20*math.log10((AFLevel_Unmod/1000)/AFLevel_Mod)
RSheet.cell(row = 2, column = 4, value = result)
print(result)# relative noise and hum level

RFile_write.save("Test_Result.xlsx")# save result

SML.close()
CMS.close()
