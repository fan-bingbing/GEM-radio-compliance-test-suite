import visa
from openpyxl import load_workbook
from openpyxl import Workbook
import re

rm = visa.ResourceManager()

SML = rm.open_resource('ASRL4::INSTR') # wanted Signal
SMB = rm.open_resource('USB0::0x0AAD::0x0054::106409::INSTR') # Unwanted Signal
#FSV = rm.open_resource('TCPIP0::192.168.10.9::hislip0::INSTR') # Spec An
CMS = rm.open_resource('GPIB0::24::INSTR') # Audio Analyzer

#scope.write_termination = '\n'
SML.clear()  # Clear instrument io buffers and status
SMB.clear()
#FSV.clear()
CMS.clear()

File_write = load_workbook(filename = "FSV_Setup.xlsx") # create a workbook from existing .xlsx file
sheet = File_write["ACS"] # load setup sheet in .xlsx to sheet

Frequency_RF = sheet["C1"].value #
Level_RF = sheet["C2"].value #
Frequency_AF = sheet["C3"].value #
Deviation = sheet["C4"].value #
Mod_state = sheet["C5"].value #
RF_power_on = sheet["C6"].value #

SML.write(f"*RST")
SML.write("SYST:DISP:UPD ON")
SML.write(f"FREQ {Frequency_RF}MHz")
SML.write(f":POW:UNIT dBuV")
SML.write(f":POW {Level_RF}dBuV")
SML.write(f":FM:INT:FREQ {Frequency_AF}kHz")
SML.write(f":FM:DEV {Deviation}kHz")
SML.write(f":FM:STAT {Mod_state}")
SML.write(f":OUTP1 {RF_power_on}")

#SML.timeout = 1000  # Acquisition timeout in milliseconds - set it higher than the sweep time
SML.query('*OPC?')

Frequency_RF = sheet["C8"].value #
Level_RF = sheet["C9"].value #
Frequency_AF = sheet["C10"].value #
Deviation = sheet["C11"].value #
Mod_state = sheet["C12"].value #
RF_power_on = sheet["C13"].value #

SMB.write(f"*RST")
#SMB.write("SYST:DISP:UPD ON")
SMB.write(f":FREQ {Frequency_RF}MHz")
SMB.write(f":UNIT:POW dBuV")
SMB.write(f":POW {Level_RF}")
SMB.write(f":FM:INT:FREQ {Frequency_AF}Hz")
SMB.write(f":FM:DEV {Deviation}kHz")
SMB.write(f":FM:STAT {Mod_state}")
SMB.write(f":OUTP1 {RF_power_on}")

#SMB.timeout = 1000  # Acquisition timeout in milliseconds - set it higher than the sweep time
SMB.query('*OPC?')


SINAD_file = Workbook()#
ws_SINAD = SINAD_file.active
ws_SINAD.title = "SINAD"

for i in range(0,8):
    SINAD_data_str = CMS.query("SINAD:R?")
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(SINAD_data_num)
    if float(SINAD_data_num) > 14.0:
        ws_SINAD.cell(row = i+1, column = 1, value = SINAD_data_num)
        Level_RF = Level_RF + 1
        SMB.write(f":POW {Level_RF}")
        SMB.query('*OPC?')
    else:
        ACS = float(SMB.query(f":POW? "))+3.5-(15.5-3.5)
        print(f"ACS result:{ACS}")

SINAD_file.save("SINAD.xlsx")

SML.close()
SMB.close()
#FSV.clear()
CMS.close()
