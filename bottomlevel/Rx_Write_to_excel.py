from openpyxl import load_workbook

def Write_to_excel(result_sheet):
    FSV_file_write = load_workbook(filename = "Test_Result.xlsx") # load an existing .xlsx file
    sheet = FSV_file_write[result_sheet] # load existing sheet named "ACP"


    
    SINAD_data_str = CMS.query("SINAD:R?")
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(SINAD_data_num)

    while float(SINAD_data_num) > 14.0:
        i=0
        sheet.cell(row = i+1, column = 1, value = SINAD_data_num)
        Level_RF = Level_RF + 1
        SMB.write(f":POW {Level_RF}")
        SMB.query('*OPC?')
        SINAD_data_str = CMS.query("SINAD:R?")
        SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
        print(SINAD_data_num)

    FSV_file_write.save("Test_Setup.xlsx") # save existing .xlsx file

    result = float(SMB.query(f":POW? "))+3.5-(15.5-3.5)
    print(f"Adjacent Channel Selectivity result:{result}")

    return result
