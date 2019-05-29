
import os
from openpyxl import Workbook # capital W

DirRoute = "C:\\Users\\afan\\Documents\\ACMA\\GX600D_VHF_maritime\\TestResult\\" # change route as required
FileName = "Tx_result.xlsx" # change file name as required


wb = Workbook()
ws = wb.active #get first active worksheet
ws.title = "active"# rename active sheet
ws1 = wb.create_sheet("Mysheet1") # insert at the end
ws2 = wb.create_sheet("Mysheet2", 0) # insert at the front

for x in range(1, 101): # row or column values must be at least 1
    for y in range(1,101):
        ws.cell(row = x, column = y) # create 100*100 cells in menory, for nothing

os.makedirs(DirRoute) # make a directory

wb.save(DirRoute + FileName) # save excel file in directory
