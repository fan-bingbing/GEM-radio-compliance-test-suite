"""
GME radio compliance test suite scripts for class and function definition .
Copyright (C) 2019 Standard Communications Pty Ltd (GME). All rights reserved.
"""

import sys
import os
import math
import visa
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import serial
import re
import config
import datetime
from decimal import * # should aviod wildcard import mentioned in PEP08
getcontext().prec = 10 # set 10 decimal values precision
# Decimal module import from decimal make sure float numbers addition yeilds correct value
import sounddevice as sd
import numpy as np
import matplotlib.pyplot as plt
from scipy.fftpack import fft
import subprocess
from os import system
from subprocess import Popen, PIPE

# below code block are for importing files from different directory
# a __init__.py file has to be exist in both "import from" and "import to" folder


class SoundCard(object):
    def __init__(self):
        self.gain = 10
        self.duration = 0.1 # sample record duration
        self.sample_rate = 44100
        self.device = ('Microphone (Sound Blaster Play!, MME',
                       'Speakers (Sound Blaster Play! 3, MME')
        sd.query_devices(self.device[0])
        sd.default.device = self.device
        sd.default.channels = 1
        sd.default.samplerate = self.sample_rate

    def get_sample(self):
        x1 = sd.rec(int(self.duration*self.sample_rate), dtype='float32', blocking=True) # record 0.1 seconds input signal from microphone
        x2 = self.gain*x1[int(self.duration*self.sample_rate/1000):int(self.duration*self.sample_rate)] # take effective samples
        x3 = x2[0:int(80*(self.sample_rate/1000))] # take first 80ms duration samples
        x4 = x3.flatten()# transfer 2D array to 1D, ready for fft operation, fft function can only take 1D array as input
        t = 80*np.linspace(0, 1, np.size(x4))# generate 80ms time axis VS samples for first 80ms

        X = fft(x4) # fft transform
        n = np.size(t)
        fr = (self.sample_rate/2)*np.linspace(0, 1, int(n/2)) # generate frequecny axis array
        X_m = (2/n)*abs(X[0:np.size(fr)])
        return {'Time_level':x4, 'Time':t, 'Freq_level':X_m, 'Freq':fr}

    def get_SINAD(self, dict): # accept returned dictionary from get_sample()
        X_mmax = np.max(dict['Freq_level']) # find maximum value in numpy array
        fr_index = np.where(dict['Freq_level'] == X_mmax) # locate index in numpy array
        fr_max = dict['Freq'][fr_index] # find corresponding frequency of maximum value

        mask = np.full(np.size(dict['Freq']), 1) # initialize a mask numpy array with size=fr, value=1

        mask[fr_index[0]]=0 # make notch filter mask by assigning some points=0 where certain bandwidth of carrier was covered
        for i in range(1,41): # notch filter cover 80 bins
            mask[fr_index[0]-i]=0
            mask[fr_index[0]+i]=0

        SND = np.sum(dict['Freq_level']) # add all points value to get SND
        ND = np.sum(dict['Freq_level'] * mask) # apply notch filter, get ND
        SINAD = 20 * np.log10(SND/ND) # calculate SINAD
        return SINAD

    def live_plots(self):
        dict = self.get_sample()
        fig, (ax0, ax1) = plt.subplots(nrows=2)
        ax0.set_title('Sinusoidal Signal')
        ax0.set_xlabel('Time(ms)')
        ax0.set_ylabel('Amplitude')
        ax0.set_ylim(-2, 2)
        line0, = ax0.plot(dict['Time'], dict['Time_level']) # time VS level plot

        ax1.set_title('Magnitude Spectrum')
        ax1.set_xlabel('Frequency(Hz)')
        ax1.set_ylabel('Magnitude')
        ax1.set_ylim(0, 1.5)
        ax3 = ax1.text(0.3, 0.9, 'SINAD', transform=ax1.transAxes)
        line1, = ax1.plot(dict['Freq'], dict['Freq_level']) # frequency VS level plot

        fig.subplots_adjust(hspace=0.6)

        while True:
            dict = self.get_sample()
            line0.set_ydata(dict['Time_level']) # update data in plots only in loop instead of update entire plots
            line1.set_ydata(dict['Freq_level'])
            ax3.set_text('SINAD = %.2f dB' % self.get_SINAD(dict=dict)) # update SINAD value

            plt.pause(0.01)# this command will call plt.show()

class SwitchBox(object):
    def __init__(self, sn, sw):
        self.sn = sn
        self.sw = sw    # A to H (depending on model)

    def Send_SW_Command(self, SW_Command):
        # print (SW_Command)
        # requires sw_usb_cs.exe and dll file in the same folder as python interpreter locate
        pipe = subprocess.Popen("sw_usb_cs.exe " + SW_Command, stdout=subprocess.PIPE)
        pipe.wait
        Sw_Reply = pipe.stdout.read()
        return Sw_Reply


    def Switch_to_Ax(self, state):
        # state is switch between output port 1 and port 2
        self.Send_SW_Command("-sn " + self.sn + " -setSPDT " + self.sw + " " + str(state))
        print(f"Set RF Switch to port{state}.")


class SpecAn(object):
    def __init__(self, address):
        self.rm = visa.ResourceManager()
        self.SP = self.rm.open_resource(address) # Spec An
        self.book = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file

    def FEP_Setup(self, freq):
        sheet = self.book["Freq_Error_Power"] # load excel sheet
        # below codes are for setting test frequency in Test_Setup.xlsx according to user's input
        sheet.cell(row = 1, column = 2, value = freq) # write test frequency
        self.book.save("Test_Setup.xlsx") # save existing .xlsx file
        # above codes are for setting test frequency in Test_Setup.xlsx according to user's input

        Centre_frequency = sheet["B1"].value # get all parameters from Test_Setup.xlsx
        Span_frequency = sheet["B2"].value
        RBW = sheet["B3"].value
        VBW = sheet["B4"].value
        RF_level = sheet["B5"].value
        Attenuation = sheet["B6"].value
        RefLev_offset = sheet["B7"].value
        Trace_peak = sheet["B8"].value
        Transducer1 = sheet["B9"].value # cable
        Trans1_ON = sheet["B10"].value
        Transducer2 = sheet["B11"].value # 30dB attenuator
        Trans2_ON = sheet["B12"].value
        Transducer3 = sheet["B13"].value # High pass filter
        Trans3_ON = sheet["B14"].value
        Limit_line_1 = sheet["B15"].value # ASNZS4365:2011
        Limit_line_1_ON = sheet["B16"].value # ASNZS4365:2011
        Limit_line_2 = sheet["B17"].value # ASNZS4295:2015
        Limit_line_2_ON = sheet["B18"].value # ASNZS4295:2015
        Sweep_points= sheet["B19"].value

        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write(f"FREQ:CENT {Centre_frequency}MHz")
        self.SP.write(f"FREQ:SPAN {Span_frequency}Hz")
        self.SP.write(f"BAND {RBW}Hz")
        self.SP.write(f"BAND:VID {VBW}Hz")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {RF_level}")
        self.SP.write(f"INP:ATT {Attenuation}")
        self.SP.write(f"{RefLev_offset}")
        self.SP.write(f"{Trace_peak}")
        self.SP.write(f"CORR:TRAN:SEL '{Transducer1}'")
        self.SP.write(f"CORR:TRAN {Trans1_ON} ")
        self.SP.write(f"CORR:TRAN:SEL '{Transducer2}'")
        self.SP.write(f"CORR:TRAN {Trans2_ON}")
        self.SP.write(f"CORR:TRAN:SEL '{Transducer3}'")
        self.SP.write(f"CORR:TRAN {Trans3_ON}")
        self.SP.write(f"CALC:LIM:NAME '{Limit_line_1}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {Limit_line_1_ON}")
        self.SP.write(f"CALC:LIM:NAME '{Limit_line_2}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {Limit_line_2_ON}")
        self.SP.write(f"SWE:POIN {Sweep_points}")

    def DeMod_Setup(self, freq):
        sheet = self.book["Analog_Demod"]
        sheet.cell(row = 1, column = 2, value = freq)
        self.book.save("Test_Setup.xlsx")

        Centre_frequency = sheet["B1"].value #get all parameters from Test_Setup.xlsx
        Dev_PerDivision = sheet["B2"].value
        Demod_BW = sheet["B3"].value
        AF_Couple = sheet["B4"].value
        RF_level = sheet["B5"].value
        Attenuation = sheet["B6"].value
        RefLev_offset = sheet["B7"].value
        Trace_Peak = sheet["B8"].value
        Demod_MT = sheet["B9"].value
        Cont_sweep = sheet["B10"].value

        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write("ADEM ON")
        self.SP.write(f"FREQ:CENT {Centre_frequency}MHz")
        self.SP.write(f"DISP:TRAC:Y:PDIV {Dev_PerDivision}kHz")
        self.SP.write(f"BAND:DEM {Demod_BW}kHz")
        self.SP.write(f"ADEM:AF:COUP {AF_Couple}")
        #Set AF coupling to AC, the frequency offset is automatically corrected.
        # i.e. the trace is always symmetric with respect to the zero line
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {RF_level}")
        self.SP.write(f"INP:ATT {Attenuation}")
        self.SP.write(f"{Trace_Peak}")
        self.SP.write(f"ADEM:MTIM {Demod_MT}ms")
        self.SP.write(f"INIT:CONT {Cont_sweep}")

    def ACP_Setup(self, freq):
        sheet = self.book["ACP"]
        sheet.cell(row = 1, column = 2, value = freq)
        self.book.save("Test_Setup.xlsx")

        Centre_frequency = sheet["B1"].value # get all parameters from Test_Setup.xlsx
        Span_frequency = sheet["B2"].value
        RBW = sheet["B3"].value
        VBW = sheet["B4"].value
        RF_level = sheet["B5"].value
        Attenuation = sheet["B6"].value
        RefLev_offset = sheet["B7"].value
        Trace_RMS = sheet["B8"].value
        Tx_CHBW = sheet["B10"].value
        AJ_CHBW = sheet["B11"].value
        AT_CHBW = sheet["B12"].value
        AJ_CHNUM = sheet["B13"].value
        AJ_SPACE = sheet["B14"].value
        AT_SPACE = sheet["B15"].value
        Power_Mode = sheet["B16"].value
        Ave_number = sheet["B17"].value

        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write("CALC:MARK:FUNC:POW:SEL ACP")
        self.SP.write(f"FREQ:CENT {Centre_frequency}MHz")
        self.SP.write(f"FREQ:SPAN {Span_frequency}kHz")
        self.SP.write(f"BAND {RBW}Hz")
        self.SP.write(f"BAND:VID {VBW}Hz")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {RF_level}")
        self.SP.write(f"INP:ATT {Attenuation}")
        self.SP.write(f"{Trace_RMS}")
        self.SP.write(f"POW:ACH:BWID:CHAN1 {Tx_CHBW}kHz")
        self.SP.write(f"POW:ACH:BWID:ACH {AJ_CHBW}kHz")
        self.SP.write(f"POW:ACH:BWID:ALT1 {AT_CHBW}kHz")
        self.SP.write(f"POW:ACH:ACP {AJ_CHNUM}")
        self.SP.write(f"POW:ACH:SPAC {AJ_SPACE}kHz")
        self.SP.write(f"POW:ACH:SPAC:ALT1 {AT_SPACE}kHz")
        self.SP.write(f"POW:ACH:MODE {Power_Mode}")
        self.SP.write(f"SWE:COUN {Ave_number}")
        self.SP.write(f"CALC:MARK:FUNC:POW:MODE WRIT")
        self.SP.write(f"DISP:TRAC:MODE AVER")

    def OBW_Setup(self, freq):
        sheet = self.book["OBW"]
        sheet.cell(row = 1, column = 2, value = freq)
        self.book.save("Test_Setup.xlsx")

        Centre_frequency = sheet["B1"].value # get all parameters from Test_Setup.xlsx
        Span_frequency = sheet["B2"].value
        RBW = sheet["B3"].value
        VBW = sheet["B4"].value
        RF_level = sheet["B5"].value
        Attenuation = sheet["B6"].value
        RefLev_offset = sheet["B7"].value
        Trace_RMS = sheet["B8"].value
        Ave_number = sheet["B9"].value

        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write("CALC:MARK:FUNC:POW:SEL OBW")
        #self.SP.write("SENS:POW:BWID:98PCT")
        self.SP.write(f"FREQ:CENT {Centre_frequency}MHz")
        self.SP.write(f"FREQ:SPAN {Span_frequency}kHz")
        self.SP.write(f"BAND {RBW}Hz")
        self.SP.write(f"BAND:VID {VBW}Hz")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {RF_level}")
        self.SP.write(f"INP:ATT {Attenuation}")
        self.SP.write(f"{Trace_RMS}")
        self.SP.write(f"SWE:COUN {Ave_number}")
        self.SP.write(f"DISP:MTAB ON")
        self.SP.write(f"CALC:MARK:FUNC:POW:MODE WRIT")
        self.SP.write(f"DISP:TRAC:MODE AVER")

    def CSE_Setup(self, sub_range, limit_line):
        if sub_range == 1:# choose sub_range
            sheet = self.book["Cond_Spurious_1"]
        elif sub_range == 2:
            sheet = self.book["Cond_Spurious_2"]
        elif sub_range == 3:
            sheet = self.book["Cond_Spurious_3"]
        elif sub_range == 4:
            sheet = self.book["Cond_Spurious_4"]
        elif sub_range == 5:
            sheet = self.book["Cond_Spurious_5"]
        else:
            raise ValueError('sub_range only accept integer 1 or 5')

        if limit_line == 2:
            sheet.cell(row = 16, column = 2, value = "ON") # choose limit line
            sheet.cell(row = 18, column = 2, value = "OFF")
            sheet.cell(row = 21, column = 2, value = "OFF")
            self.book.save("Test_Setup.xlsx")
        elif limit_line == 3:
            sheet.cell(row = 16, column = 2, value = "OFF") # choose limit line
            sheet.cell(row = 18, column = 2, value = "ON")
            sheet.cell(row = 21, column = 2, value = "OFF")
            self.book.save("Test_Setup.xlsx")
        elif limit_line == 4:
            sheet.cell(row = 16, column = 2, value = "OFF") # choose limit line
            sheet.cell(row = 18, column = 2, value = "OFF")
            sheet.cell(row = 21, column = 2, value = "ON")
            self.book.save("Test_Setup.xlsx")
        else:
            raise ValueError('limit_line only accept integer 2, 3 or 4')

        start_frequency = sheet["B1"].value # get all parameters from Test_Setup.xlsx
        stop_frequency = sheet["B2"].value
        RBW = sheet["B3"].value
        VBW = sheet["B4"].value
        RF_level = sheet["B5"].value
        Attenuation = sheet["B6"].value
        RefLev_offset = sheet["B7"].value
        Trace_peak = sheet["B8"].value
        Transducer1 = sheet["B9"].value
        Trans1_ON = sheet["B10"].value
        Transducer2 = sheet["B11"].value
        Trans2_ON = sheet["B12"].value
        Transducer3 = sheet["B13"].value
        Trans3_ON = sheet["B14"].value
        Limit_line_1 = sheet["B15"].value
        Limit_line_1_ON = sheet["B16"].value
        Limit_line_2 = sheet["B17"].value
        Limit_line_2_ON = sheet["B18"].value
        Sweep_points= sheet["B19"].value
        Limit_line_3 = sheet["B20"].value
        Limit_line_3_ON = sheet["B21"].value



        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write(f"FREQ:STAR {start_frequency}MHz")
        self.SP.write(f"FREQ:STOP {stop_frequency}MHz")
        self.SP.write(f"BAND {RBW}kHz")
        self.SP.write(f"BAND:VID {VBW}kHz")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {RF_level}")
        self.SP.write(f"INP:ATT {Attenuation}")
        self.SP.write(f"{RefLev_offset}")
        self.SP.write(f"{Trace_peak}")
        self.SP.write(f"CORR:TRAN:SEL '{Transducer1}'")
        self.SP.write(f"CORR:TRAN {Trans1_ON} ")
        self.SP.write(f"CORR:TRAN:SEL '{Transducer2}'")
        self.SP.write(f"CORR:TRAN {Trans2_ON}")
        self.SP.write(f"CORR:TRAN:SEL '{Transducer3}'")
        self.SP.write(f"CORR:TRAN {Trans3_ON}")
        self.SP.write(f"CALC:LIM:NAME '{Limit_line_1}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {Limit_line_1_ON}")
        self.SP.write(f"CALC:LIM:NAME '{Limit_line_2}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {Limit_line_2_ON}")
        self.SP.write(f"SWE:POIN {Sweep_points}")
        self.SP.write(f"CALC:LIM:NAME '{Limit_line_3}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {Limit_line_3_ON}")
        self.SP.write(f"DISP:TRAC:MODE MAXH")

    def TranP_Setup(self, freq):
        sheet = self.book["Tran_Perform"]
        sheet.cell(row = 1, column = 2, value = freq)
        self.book.save("Test_Setup.xlsx")

        Centre_frequency = sheet["B1"].value # get all parameters from Test_Setup.xlsx
        Dev_PerDivision = sheet["B2"].value
        Demod_BW = sheet["B3"].value
        AF_Couple = sheet["B4"].value
        RF_level = sheet["B5"].value
        Attenuation = sheet["B6"].value
        RefLev_offset = sheet["B7"].value
        Trace_Peak = sheet["B8"].value
        Demod_MT = sheet["B9"].value
        Cont_sweep = sheet["B10"].value
        Trigger_offset = sheet["B11"].value

        self.SP.write(f"*RST") # write all paramaters to SpecAn
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write("ADEM ON")
        self.SP.write(f"FREQ:CENT {Centre_frequency}MHz")
        self.SP.write(f"DISP:TRAC:Y:PDIV {Dev_PerDivision}kHz")
        self.SP.write(f"BAND:DEM {Demod_BW}kHz")
        self.SP.write(f"ADEM:AF:COUP {AF_Couple}")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {RefLev_offset}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {RF_level}")
        self.SP.write(f"INP:ATT {Attenuation}")
        self.SP.write(f"{Trace_Peak}")
        self.SP.write(f"ADEM:MTIM {Demod_MT}ms")
        self.SP.write(f"INIT:CONT {Cont_sweep}")
        self.SP.write(f"TRIG:HOLD {Trigger_offset}ms")

    def write(self, str):
        self.SP.write(str)

    def query(self, str):
        return self.SP.query(str)

    def get_FEP_result(self, freq):
        self.SP.write("CALC:MARK1:MAX")
        Frequency = self.SP.query("CALC:MARK1:X?")
        Level = float(self.SP.query("CALC:MARK1:Y?"))
        Frequency_error = float(Frequency) - float(self.SP.query("FREQ:CENT?"))
        indication = (self.SP.query("*OPC?")).replace("1","Completed.")
        self.screenshot('FEP_'+str(freq)+'MHz')
        return {'F':Frequency_error, 'P':Level, 'I':indication}

    def get_OBW_result(self, channel):
        OBW_str = self.SP.query("CALC:MARK:FUNC:POW:RES? AOBW")
        OBS_list = re.findall(r'\d+\.\d+', OBW_str )
        frequency_centre = float(self.SP.query("FREQ:CENT?"))*1e-6
        frequency_low = float(OBS_list[1])*1e-6
        frequency_high = float(OBS_list[3])*1e-6
        OBW = float(OBS_list[0])*1e-3
        indication = (self.SP.query("*OPC?")).replace("1","Completed.")
        self.screenshot('OBW_'+str(channel))
        #print(frequency_centre)
        #print (OBW_str)
        #print (OBS_list)
        return {'F0':frequency_centre, 'F1':frequency_low,
                'F2':frequency_high, 'OBW':OBW, 'I':indication}

    def get_CSE_result(self):
        self.SP.write("CALC:MARK1:MAX")
        self.SP.write("CALC:MARK2:MAX")
        self.SP.write("CALC:MARK2:MAX:NEXT")
        Frequency1 = float(self.SP.query("CALC:MARK1:X?"))/1e6
        Level1 = float(self.SP.query("CALC:MARK1:Y?"))

        Frequency2 = float(self.SP.query("CALC:MARK2:X?"))/1e6
        Level2 = float(self.SP.query("CALC:MARK2:Y?"))

        indication = (self.SP.query("*OPC?")).replace("1","Completed.")
        return {'F1':Frequency1, 'P1':Level1, 'F2':Frequency2, 'P2':Level2, 'I':indication}

    def screenshot(self, file_name):
        self.SP.write("HCOP:DEV:LANG PNG") # set file type to .png
        self.SP.write("HCOP:CMAP:DEF4")
        self.SP.write(f"MMEM:NAME \'c:\\temp\\Dev_Screenshot.png\'")
        self.SP.write("HCOP:IMM") # perform copy and save .png file on SpecAn
        self.SP.query("*OPC?")

        file_data = self.SP.query_binary_values(f"MMEM:DATA? \'c:\\temp\\Dev_Screenshot.png\'", datatype='s',)[0] # query binary data and save
        new_file = open(f"c:\\Temp\\{file_name}.png", "wb")# open a new file as "binary/write" on PC
        new_file.write(file_data) # copy data to the file on PC
        new_file.close()
        print(f"Screenshot saved to PC c:\\Temp\\{file_name}.png\n")

    def close(self):
        self.SP.close()

class SigGen(object):

    def __init__(self, address):
        self.rm = visa.ResourceManager()
        self.SG = self.rm.open_resource(address)
        self.book = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file


    def Tx_Setup(self):
        sheet = self.book["Tx_Siggen_Setting"]

        Frequency_AF = sheet["C1"].value # get all parameters from Test_Setup.xlsx
        self.Level_AF = sheet["C2"].value
        AF_output_on = sheet["C3"].value

        self.SG.write(f"*RST") # write all paramaters to SigGen
        self.SG.write("SYST:DISP:UPD ON")
        #self.SG.write(f":FM:INT:FREQ {Frequency_AF}kHz")
        self.SG.write(f"LFO:FREQ {Frequency_AF}kHz")
        # self.SG.write(f":OUTP2:VOLT {self.Level_AF}mV")
        self.SG.write(f"LFO:VOLT {self.Level_AF}mV")
        # self.SG.write(f":OUTP2 {AF_output_on}")
        self.SG.write(f"LFO {AF_output_on}")

    def TranP_Setup(self, freq):
        sheet = self.book["Tran_Perform"]
        sheet.cell(row = 1, column = 7, value = freq)
        self.book.save("Test_Setup.xlsx")


        Frequency_RF = sheet["G1"].value # get all parameters from Test_Setup.xlsx
        self.Level_RF = sheet["G2"].value
        Frequency_AF = sheet["G3"].value
        Deviation = sheet["G4"].value
        Mod_state = sheet["G5"].value
        RF_power_on = sheet["G6"].value

        self.SG.write(f"*RST") # write all paramaters to SigGen
        self.SG.write("SYST:DISP:UPD ON")
        self.SG.write(f"FREQ {Frequency_RF}MHz")
        self.SG.write(f":POW:UNIT dBuV")
        self.SG.write(f":POW {self.Level_RF}dBuV")
        self.SG.write(f"LFO:FREQ {Frequency_AF}kHz")
        self.SG.write(f"FM {Deviation}kHz")
        self.SG.write(f"FM:STAT {Mod_state}")
        self.SG.write(f"OUTP {RF_power_on}")

    def Unwanted_Signal(self, freq, unwanted_level):
        sheet = self.book["Rx_Siggen_Setting"]
        sheet.cell(row = 8, column = 3, value = freq)
        sheet.cell(row = 9, column = 3, value = unwanted_level)
        self.book.save("Test_Setup.xlsx")

        Frequency_RF = sheet["C8"].value # get all parameters from Test_Setup.xlsx
        self.Level_RF = sheet["C9"].value
        Frequency_AF = sheet["C10"].value
        Deviation = sheet["C11"].value
        Mod_state = sheet["C12"].value
        RF_power_on = sheet["C13"].value

        self.SG.write(f"*RST") # write all paramaters to SigGen
        self.SG.write("SYST:DISP:UPD ON")
        self.SG.write(f"FREQ {Frequency_RF}MHz")
        self.SG.write(f":UNIT:POW dBuV")
        self.SG.write(f":POW {self.Level_RF}dBuV")
        self.SG.write(f":FM:INT:FREQ {Frequency_AF}Hz")
        self.SG.write(f"FM {Deviation}kHz")
        self.SG.write(f":FM:STAT {Mod_state}")
        self.SG.write(f":OUTP1 {RF_power_on}")

    def Wanted_Signal(self, freq):
        sheet = self.book["Rx_Siggen_Setting"]
        sheet.cell(row = 1, column = 3, value = freq)
        self.book.save("Test_Setup.xlsx")

        Frequency_RF = sheet["C1"].value # get all parameters from Test_Setup.xlsx
        self.Level_RF = sheet["C2"].value
        Frequency_AF = sheet["C3"].value
        Deviation = sheet["C4"].value
        Mod_state = sheet["C5"].value
        RF_power_on = sheet["C6"].value

        self.SG.write(f"*RST") # write all paramaters to SigGen
        self.SG.write("SYST:DISP:UPD ON")
        self.SG.write(f"FREQ {Frequency_RF}MHz")
        self.SG.write(f":POW:UNIT dBuV")
        self.SG.write(f":POW {self.Level_RF}dBuV")
        self.SG.write(f"LFO:FREQ {Frequency_AF}kHz")
        self.SG.write(f"FM {Deviation}kHz")
        self.SG.write(f"FM:STAT {Mod_state}")
        self.SG.write(f"OUTP {RF_power_on}")

    def Set_Timeout(self, ms):
        self.SG.timeout = ms # useful on CMS for Rx test, pyvisa parameter

    def Lev_AF(self):
        return self.Level_AF # useful for Max_deviation test

    def Lev_RF(self):
        return self.Level_RF # useful for Rx test

    def write(self, str):
        self.SG.write(str)

    def query(self, str):
        return self.SG.query(str)

    def close(self):
        self.SG.close()


class Radio(object):
    def __init__(self, com):
        self.ESC_CHAR = 0x7D # Escape char
        self.STX_CHAR =  0x7E # Start of packet
        self.ETX_CHAR = 0x7F # End of packet
        self.CHECKSUM_XOR_MASK = 0xFF

        self.payload0 = bytearray(b'\xB4\x88\x2a\x80')# set carrie freq 136MHz basd on 12.5kHz calculation
        self.payload1 = bytearray(b'\xB4\x93\x03')# set 25W power
        self.payload2 = bytearray(b'\xA6\x01') # PTT ON
        self.payload3 = bytearray(b'\xB4\x91\x03\xE8')# 1kHz audio on
        self.payload4 = bytearray(b'\xB4\x91\x0B\xB8')# 3kHz audio on
        self.payload5 = bytearray(b'\xB4\x91\x00\x00')# audio off
        self.payload6 = bytearray(b'\xB4\x82\x10')# selcall tone on
        self.payload7 = bytearray(b'\xB4\x82\x0f')# selcall tone off
        self.payload8 = bytearray(b'\xA6\x00') # PTT off

        self.port = com
        self.baudrate = 115200
        self.timeout = None
        self.ser = serial.Serial(port=self.port, baudrate=self.baudrate, timeout=self.timeout)  # open serial port

    def Packet_Gen(self, payload): # return GME2 protocol packet
        # Header
        tx_array = bytearray((self.STX_CHAR, 2, len(payload)))

        # Escape the payload, append it to the output buffer, make the checksum
        chksum = 0
        for i, b in enumerate(payload):
            # The first word of the payload is always the command ID.
            # The MSB of the command ID is always 1. The radio just
            # clears it in the ACK, without doing anything else with it.
            if i == 0:
                bb = (0x80 | b) & 0xFF
            else:
                bb = b & 0xFF
            chksum ^= bb
            if bb in (self.ESC_CHAR, self.STX_CHAR, self.ETX_CHAR):
                tx_array.append(self.ESC_CHAR)
                tx_array.append(bb ^ 0x20)
            else:
                tx_array.append(bb)

        # Trailer
        tx_array.append(chksum ^ self.CHECKSUM_XOR_MASK)
        tx_array.append(self.ETX_CHAR)

        return tx_array

    def Set_Freq(self, freq):
        a = ((hex(int((Decimal(freq)*Decimal(1e6))/(Decimal(12.5*1e3))))[2]+hex(int((Decimal(freq)*Decimal(1e6))/Decimal((12.5*1e3))))[3]))# calculate first HEX of input frequency
        b = ((hex(int((Decimal(freq)*Decimal(1e6))/(Decimal(12.5*1e3))))[4]+hex(int((Decimal(freq)*Decimal(1e6))/Decimal((12.5*1e3))))[5]))# calculate second HEX of input frequency
        self.payload0[2] = int(a,16) # assign first DEC (transferred from HEX) to the third number of paylaod0
        self.payload0[3] = int(b,16) # assign second DEC (transferred from HEX) to the fourth number of paylaod0
        self.ser.write(self.Packet_Gen(self.payload0))
        print(f"radio frequency has been set to {freq} MHz.")

    def Set_Pow(self, pow):
        if pow == "low":
            self.payload1[2] = 0
        elif pow == "high":
            self.payload1[2] = 3
        else:
            print("Set_Pow() only accept 'low' or 'high'")
        print(f"radio power has been set to {pow}.")

        self.ser.write(self.Packet_Gen(self.payload1))

    def Radio_On(self):
        self.ser.write(self.Packet_Gen(self.payload2))
        print("Radio's on")

    def Radio_Off(self):
        self.ser.write(self.Packet_Gen(self.payload8))
        print("Radio's off")

    def Radio_close(self):
        self.ser.close()
        print("Serial session is closed.")

class Excel(object):
    def __init__(self, file_name):
        self.file = load_workbook(filename = file_name) # load Test_Result.xlsx

    def get_sheet(self, sheet_name):
        self.sheet = self.file[sheet_name]

    def write(self, row, column, value):
        self.sheet.cell(row = row, column = column, value = value)

    def clear(self, start_cell, end_cell):
        for row in self.sheet[start_cell +':'+ end_cell]:# clear certain block of cells in selected sheet
            for cell in row:
                cell.value = None

    def save(self, file_name):
        self.file.save(file_name)


def copy_excel(to_file, sheet_list): # copy multiple sheets from Test_Result.xlsx to a new excel file in desktop folder
    wb = openpyxl.Workbook()

    wb.save(f'C:\\Users\\Compliance_test\\Desktop\\exp\\{to_file}.xlsx')

    with pd.ExcelWriter(f'C:\\Users\\Compliance_test\\Desktop\\exp\\{to_file}.xlsx') as writer:  # doctest: +SKIP

        for item in sheet_list: # varable item get whatever stored in sheet_list directly
            data = pd.read_excel(r'C:\documents\Aaron_test\aaron test\compliance test suite\Test_Result.xlsx', sheet_name=item)
            data.to_excel(writer, sheet_name=item)

    print("transfer completed.")

# example usage: sheet_list = ['Ferror_Pow', 'ACP', 'Cond_Spur_4268' ]
# example usage: copy_excel('haha', sheet_list)



def Tx_set_standard_test_condition():
    Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0 #get the initial deviation value
    Level_AF = SML.Lev_AF()# initial Level_AF
# above code is to find Audio output level
# satisfying standard condtion (around 1.5kHz deviation)
    if Dev_Reading < 1.5:
        while Dev_Reading < 1.47:
            print(f"current deviation:{Dev_Reading}kHz")
            Level_AF = Level_AF+1
            SML.write(f"LFO:VOLT {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0
    else:
        while Dev_Reading > 1.53:
            print(f"current deviation:{Dev_Reading}kHz")
            Level_AF = Level_AF-1
            SML.write(f"LFO:VOLT  {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0

    FSV.write(f"INIT:CONT ON")
    FSV.query('*OPC?')
    print(f"Audio level has been set to {Level_AF} mV")
    return Level_AF

def Tx_Frequency_error_Carrier_power(freq):

    # Result_sheet.get_sheet("Ferror_Pow")
    FSV.FEP_Setup(freq)
    FSV.query('*OPC?')
    CP50.Set_Freq(freq)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(2)
    FSV.write("DISP:TRAC:MODE MAXH")
    time.sleep(3)
    FSV.write("DISP:TRAC:MODE VIEW")
    CP50.Radio_Off()
    FSV.screenshot('FEP_' + str(freq) + 'MHz')
    FEP = FSV.get_FEP_result(freq)
    print(f"Frequency error:{FEP['F']}Hz")
    print(f"Carrier power:{FEP['P']}dBm")
    return FEP

def CHSW_FEP(start_f, start_row, end_row, clear):
    Start_F = Decimal(start_f)/Decimal(1)
    Result_sheet.get_sheet("Ferror_Pow")
    if clear == 1:
        Result_sheet.clear('A2', 'D15')
    else:
        pass
    for i in range(start_row, end_row): # run whatever number of nchannels with 12.5kHz channel spacing
        FEP = Tx_Frequency_error_Carrier_power(freq=Decimal(Start_F))

        Result_sheet.write(row = i+2, column = 1, value = Start_F)
        Result_sheet.write(row = i+2, column = 2, value = FEP['F'])
        Result_sheet.write(row = i+2, column = 3, value = FEP['P'])
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 4, value = Timestamp)
        Result_sheet.save("Test_Result.xlsx")
        Start_F += Decimal(0.0125) # frequency in MHz
    print(f"Frequency error and Carrier power test {FEP['I']}")

def Tx_Max_deviation(freq, sheet):

    AF_list1 = [100, 300, 500, 700, 900, 1000, 1300, 1500, 1700, 1900, 2000, 2300, 2550] # audio frequency list in Hz
    AF_list2 = [3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000, 11000, 12000, 12500]
    Reading_array1 = np.zeros(shape=(13,3)) # empty numpy array for below 3kHz result storage
    Reading_array2 = np.zeros(shape=(11,3)) # empty numpy array for above 3kHz result storage
    # RSheet = Get_result_sheet("Test_Result.xlsx", "Max_Dev")
    Result_sheet.get_sheet(sheet)
    Result_sheet.clear('A2', 'F26')
    SML.Tx_Setup()
    SML.query('*OPC?')
    FSV.DeMod_Setup(freq)
    FSV.query('*OPC?')
    CP50.Set_Freq(freq)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(2)
    FSV.write(f"INIT:CONT OFF")

    Level_AF = Tx_set_standard_test_condition()

    # below code block is to vary audio frequency to complete the test
    Level_AF = 10*Level_AF # bring audio level up 20dB in one step( 10times valtage)
    SML.write(f"LFO:VOLT {Level_AF}mV")
    print(f"Audio level has been set to {Level_AF} mV")

    for i in range(0,13):
        print(f"At Audio frequency:{AF_list1[i]}")
        SML.write(f"LFO:FREQ {AF_list1[i]}Hz")
        SML.query('*OPC?')
        time.sleep(2)
        FSV.write(f"INIT:CONT OFF")
        time.sleep(1)
        Reading_array1[i][0]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? PPE"))/1000
        Reading_array1[i][1]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MPE"))/1000
        Reading_array1[i][2]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000

        FSV.query("*OPC?")
        FSV.write(f"INIT:CONT ON")
        print(f"Deviation is {Reading_array1[i]}kHz")
        Result_sheet.write(row = i+2, column = 1, value = freq)
        Result_sheet.write(row = i+2, column = 2, value = AF_list1[i])
        Result_sheet.write(row = i+2, column = 3, value = Reading_array1[i][0])
        Result_sheet.write(row = i+2, column = 4, value = Reading_array1[i][1])
        Result_sheet.write(row = i+2, column = 5, value = Reading_array1[i][2])
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 6, value = Timestamp)

    Result_sheet.save("Test_Result.xlsx")

    Level_AF = Level_AF/10
    SML.write(f"LFO:VOLT {Level_AF}mV")
    print(f"Audio level has been set to {Level_AF} mV")

    for i in range(0,11):
        print(f"At Audio frequency:{AF_list2[i]}")
        SML.write(f"LFO:FREQ {AF_list2[i]}Hz")
        SML.query('*OPC?')
        time.sleep(1)
        FSV.write(f"INIT:CONT OFF")
        time.sleep(1)
        Reading_array2[i][0]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? PPE"))/1000
        Reading_array2[i][1]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MPE"))/1000
        Reading_array2[i][2]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000

        FSV.query("*OPC?")
        FSV.write(f"INIT:CONT ON")
        print(f"Deviation is {Reading_array2[i]}kHz")
        Result_sheet.write(row = i+16, column = 1, value = freq)
        Result_sheet.write(row = i+16, column = 2, value = AF_list2[i])
        Result_sheet.write(row = i+16, column = 3, value = Reading_array2[i][0])
        Result_sheet.write(row = i+16, column = 4, value = Reading_array2[i][1])
        Result_sheet.write(row = i+16, column = 5, value = Reading_array2[i][2])
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+16, column = 6, value = Timestamp)

    Result_sheet.save("Test_Result.xlsx")
    CP50.Radio_Off()
    SML.write("LFO OFF")# turn off audio output at the end of the test
    indication = (FSV.query("*OPC?")).replace("1","Completed.")
    print(f"Maximum Deviation test {indication}")

def Tx_Max_deviation_CB(freq, sheet):

    AF_list = [300, 500, 700, 900, 1100, 1300, 1500, 1700, 1900, 2100, 2300, 2550]
    Reading_array = np.zeros(shape=(20, 3))
    Result_sheet.get_sheet(sheet)
    Result_sheet.clear('A2', 'F17')

    SML.Tx_Setup()
    SML.query('*OPC?')
    FSV.DeMod_Setup(freq)
    FSV.query('*OPC?')
    CP50.Set_Freq(freq)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(2)
    FSV.write(f"INIT:CONT OFF")

    Level_AF = Tx_set_standard_test_condition()

    time.sleep(2)
    FSV.write("TRIG:SOUR EXT") # set to EXT trigger
    FSV.write("TRIG:LEV 0.5V") # set to EXT trigger level to 500mV
    time.sleep(2)
    FSV.write("INIT:CONT OFF") # set to single sweep
    time.sleep(2)
    FSV.write("INIT:CONM") # set to wait for trigger
    time.sleep(2)
    Level_AF = 100*Level_AF # bring audio level up 20dB in one step
    SML.write(f"LFO:VOLT {Level_AF}mV")
    time.sleep(2)
    print(f"Audio level has been set to {Level_AF} mV")

    Reading_array[0][0]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? PPE"))/1000
    Reading_array[0][1]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MPE"))/1000
    Reading_array[0][2]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000
    FSV.screenshot('1kHz_INS_deviation_'+str(freq)+'MHz')
    Result_sheet.write(row = 2, column = 1, value = freq)
    Result_sheet.write(row = 2, column = 2, value = 1000)
    Result_sheet.write(row = 2, column = 3, value = Reading_array[0][0])
    Result_sheet.write(row = 2, column = 4, value = Reading_array[0][1])
    Result_sheet.write(row = 2, column = 5, value = Reading_array[0][2])
    Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
    Result_sheet.write(row = 2, column = 6, value = Timestamp)
    FSV.write("TRIG:SOUR IMM")
    FSV.write(f"INIT:CONT ON")

    time.sleep(2)
    FSV.write(f"INIT:CONT OFF")
    Reading_array[1][0]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? PPE"))/1000
    Reading_array[1][1]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MPE"))/1000
    Reading_array[1][2]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000
    FSV.screenshot('1kHz_SSD_deviation_'+str(freq)+'MHz')
    Result_sheet.write(row = 3, column = 1, value = freq)
    Result_sheet.write(row = 3, column = 2, value = 1000)
    Result_sheet.write(row = 3, column = 3, value = Reading_array[1][0])
    Result_sheet.write(row = 3, column = 4, value = Reading_array[1][1])
    Result_sheet.write(row = 3, column = 5, value = Reading_array[1][2])
    Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
    Result_sheet.write(row = 3, column = 6, value = Timestamp)
    FSV.write(f"INIT:CONT ON")

    Level_AF = Level_AF/100 # bring audio level up 20dB in one step
    SML.write(f"LFO:VOLT {Level_AF}mV")
    print(f"Audio level has been set to {Level_AF} mV")

    CP50.Radio_Off() # to start another 1min Tx
    time.sleep(1)
    CP50.Radio_On()

    for i in range(0,12):
        print(f"At Audio frequency:{AF_list[i]}")
        SML.write(f"LFO:FREQ {AF_list[i]}Hz")
        SML.query('*OPC?')
        time.sleep(1)
        FSV.write(f"INIT:CONT OFF")
        time.sleep(1)
        Reading_array[i+2][0]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? PPE"))/1000
        Reading_array[i+2][1]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MPE"))/1000
        Reading_array[i+2][2]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000

        FSV.query("*OPC?")
        FSV.write(f"INIT:CONT ON")
        print(f"Deviation is {Reading_array[i+2]}kHz")
        Result_sheet.write(row = i+4, column = 1, value = freq)
        Result_sheet.write(row = i+4, column = 2, value = AF_list[i])
        Result_sheet.write(row = i+4, column = 3, value = Reading_array[i+2][0])
        Result_sheet.write(row = i+4, column = 4, value = Reading_array[i+2][1])
        Result_sheet.write(row = i+4, column = 5, value = Reading_array[i+2][2])
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+4, column = 6, value = Timestamp)

    max = np.max(Reading_array[2:, 2]) # find maximum value in third column excluding first two rows
    index = np.where(Reading_array[2:, 2] == max) # locate index in this column
    # print(max)
    # print(index)
    # print(index[0])
    # print(index[0][0])
    Audio_F = AF_list[int(index[0][0])] # find corresponding audio frequency of maximum value
    print(f"Maximum deviation occurs at {Audio_F} Hz")

    SML.query('*OPC?')
    SML.write(f"LFO:FREQ {Audio_F}Hz")
    print(f"Audio freqency has been set to {Audio_F} Hz")
    time.sleep(2)
    FSV.write("TRIG:SOUR EXT") # set to EXT trigger
    FSV.write("TRIG:LEV 0.5V") # set to EXT trigger level to 500mV
    time.sleep(2)
    FSV.write("INIT:CONT OFF") # set to single sweep
    time.sleep(2)
    FSV.write("INIT:CONM") # set to wait for trigger
    time.sleep(2)
    Level_AF = 100*Level_AF # bring audio level up 20dB in one step
    SML.write(f"LFO:VOLT {Level_AF}mV")
    time.sleep(2)
    print(f"Audio level has been set to {Level_AF} mV")

    Reading_array[14][0]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? PPE"))/1000
    Reading_array[14][1]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MPE"))/1000
    Reading_array[14][2]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000
    FSV.screenshot(f'{Audio_F}Hz_INS_deviation_'+str(freq)+'MHz')
    Result_sheet.write(row = 16, column = 1, value = freq)
    Result_sheet.write(row = 16, column = 2, value = Audio_F)
    Result_sheet.write(row = 16, column = 3, value = Reading_array[14][0])
    Result_sheet.write(row = 16, column = 4, value = Reading_array[14][1])
    Result_sheet.write(row = 16, column = 5, value = Reading_array[14][2])
    Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
    Result_sheet.write(row = 16, column = 6, value = Timestamp)
    FSV.write("TRIG:SOUR IMM")
    FSV.write(f"INIT:CONT ON")

    time.sleep(2)
    FSV.write(f"INIT:CONT OFF")
    Reading_array[15][0]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? PPE"))/1000
    Reading_array[15][1]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MPE"))/1000
    Reading_array[15][2]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000
    FSV.screenshot(f'{Audio_F}Hz_SSD_deviation_'+str(freq)+'MHz')
    Result_sheet.write(row = 17, column = 1, value = freq)
    Result_sheet.write(row = 17, column = 2, value = Audio_F)
    Result_sheet.write(row = 17, column = 3, value = Reading_array[15][0])
    Result_sheet.write(row = 17, column = 4, value = Reading_array[15][1])
    Result_sheet.write(row = 17, column = 5, value = Reading_array[15][2])
    Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
    Result_sheet.write(row = 17, column = 6, value = Timestamp)
    FSV.write(f"INIT:CONT ON")


    SML.write(f"LFO OFF")
    CP50.Radio_Off()

    Result_sheet.save("Test_Result.xlsx")
    print(f"Maximum Deviation test for CB completed.")


def Tx_Out_of_band_modulation_response(freq, sheet):

    AF_list = [1000, 2550, 3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000, 11000, 12000, 12500]
    Reading_array = np.zeros(shape=(13,3))
    Result_sheet.get_sheet(sheet)
    Result_sheet.clear('A2', 'F14')
    SML.Tx_Setup()
    SML.query('*OPC?')
    FSV.DeMod_Setup(freq)
    FSV.query('*OPC?')
    CP50.Set_Freq(freq)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(2)
    FSV.write(f"INIT:CONT OFF")

    Tx_set_standard_test_condition()

    for i in range(0,13):
        print(f"At Audio frequency:{AF_list[i]}")
        SML.write(f"LFO:FREQ {AF_list[i]}Hz")
        SML.query('*OPC?')
        time.sleep(1)
        FSV.write(f"INIT:CONT OFF")
        time.sleep(1)
        Reading_array[i][0]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? PPE"))/1000
        Reading_array[i][1]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MPE"))/1000
        Reading_array[i][2]= float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000

        FSV.query("*OPC?")
        FSV.write(f"INIT:CONT ON")
        print(f"Deviation is {Reading_array[i]}kHz")
        Result_sheet.write(row = i+2, column = 1, value = freq)
        Result_sheet.write(row = i+2, column = 2, value = AF_list[i])
        Result_sheet.write(row = i+2, column = 3, value = Reading_array[i][0])
        Result_sheet.write(row = i+2, column = 4, value = Reading_array[i][1])
        Result_sheet.write(row = i+2, column = 5, value = Reading_array[i][2])
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 6, value = Timestamp)

    Result_sheet.save("Test_Result.xlsx")
    CP50.Radio_Off()
    SML.write("LFO OFF")# turn off audio output at the end of the test
    indication = (FSV.query("*OPC?")).replace("1","Completed.")
    print(f"Out of Band Response test {indication}")


def Tx_Adjacent_channel_power(freq):

    Result_sheet.get_sheet("ACP")
    SML.Tx_Setup()
    SML.query('*OPC?')
    FSV.DeMod_Setup(freq)
    FSV.query('*OPC?')
    CP50.Set_Freq(freq)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(2)
    FSV.write(f"INIT:CONT OFF")

    Tx_set_standard_test_condition()

    FSV.ACP_Setup(freq)
    time.sleep(8)
    FSV.write(f"DISP:TRAC:MODE VIEW")
    ACP = FSV.query("CALC:MARK:FUNC:POW:RES? ACP")
    print(ACP)
    FSV.screenshot('ACP_'+str(freq)+'MHz')
    ACP_LIST = re.findall(r'-?\d+\.\d+', ACP) # -? with or without negative sign, \d+ one or more digit
    CP50.Radio_Off()
    SML.write("LFO OFF")# turn off audio output at the end of the test
    # print(ACP_LIST)
    return ACP_LIST

def CHSW_ACP(start_f, start_row, end_row, clear):
    Start_F = Decimal(start_f)/Decimal(1)
    Result_sheet.get_sheet("ACP")
    if clear == 1:
        Result_sheet.clear('A2', 'E15')
    else:
        pass
    for i in range(start_row, end_row): # run whatever number of nchannels with 12.5kHz channel spacing
        ACP_LIST = Tx_Adjacent_channel_power(freq=Decimal(Start_F))
        Result_sheet.write(row = i+2, column = 1, value = Start_F)
        Result_sheet.write(row = i+2, column = 2, value = ACP_LIST[0])
        Result_sheet.write(row = i+2, column = 3, value = ACP_LIST[1])
        Result_sheet.write(row = i+2, column = 4, value = ACP_LIST[2])
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 5, value = Timestamp)
        Result_sheet.save("Test_Result.xlsx")
        Start_F += Decimal(0.0125) # frequency in MHz
        indication = (FSV.query("*OPC?")).replace("1","Completed.")
    print(f"Adjacent channel power test {indication}")

def Tx_Occupied_bandwidth(freq, channel):

    Result_sheet.get_sheet("OBW")
    SML.Tx_Setup()
    SML.query('*OPC?')
    FSV.DeMod_Setup(freq=freq)
    FSV.query('*OPC?')
    CP50.Set_Freq(freq=freq)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(2)
    FSV.write(f"INIT:CONT OFF")

    Tx_set_standard_test_condition()

    FSV.OBW_Setup(freq=freq)
    FSV.query('*OPC?')
    time.sleep(8)
    FSV.write(f"DISP:TRAC:MODE VIEW")
    if channel == 'top':
        dict = FSV.get_OBW_result(channel=channel)
        Result_sheet.write(row = 2, column = 1, value = dict['F0'])
        Result_sheet.write(row = 2, column = 2, value = dict['F1'])
        Result_sheet.write(row = 2, column = 3, value = dict['F2'])
        Result_sheet.write(row = 2, column = 4, value = dict['OBW'])
    else:
        dict = FSV.get_OBW_result(channel=channel)
        Result_sheet.write(row = 3, column = 1, value = dict['F0'])
        Result_sheet.write(row = 3, column = 2, value = dict['F1'])
        Result_sheet.write(row = 3, column = 3, value = dict['F2'])
        Result_sheet.write(row = 3, column = 4, value = dict['OBW'])

    Result_sheet.save("Test_Result.xlsx") # save existing .xlsx file
    CP50.Radio_Off()
    SML.write("LFO OFF")# turn off audio output at the end of the test
    print(dict)

def Tx_Conducted_spurious_emissions(freq, sheet, clear):

    while True:
        print("1. low subranges test without high pass filter, make sure no filter installed:")
        print("2. high subranges test with high pass filter, make sure filter installed:")
        print("3. go back:")
        choice = input("> ")
        if choice == "1":
            for i in range(0, 3):# first 3 subranges without high pass filter
                FSV.CSE_Setup(sub_range=i+1, limit_line=config.limit_line_factor)
                FSV.query('*OPC?')
                CSE_operation(freq=freq, row=i+2, sub_range=i+1, sheet=sheet, clear=clear)
                FSV.screenshot('CSE0'+str(i+1)+'_'+str(freq)+'MHz')

        elif choice == "2":
            for i in range(3, 5):# second 2 subranges with high pass filter
                FSV.CSE_Setup(sub_range=i+1, limit_line=config.limit_line_factor)
                FSV.query('*OPC?')
                CSE_operation(freq=freq, row=i+2, sub_range=i+1, sheet=sheet, clear=clear)
                FSV.screenshot('CSE0'+str(i+1)+'_'+str(freq)+'MHz')

        elif choice == "3":
            break

        else:
            print("choice only take number 1, 2 or 3, try agin...")

def CSE_operation(freq, row, sub_range, sheet, clear):
    Result_sheet.get_sheet(sheet)
    if clear == 1:
        Result_sheet.clear('A2', 'G6')
    else:
        pass
    CP50.Set_Freq(freq)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(5)
    FSV.write("DISP:TRAC:MODE VIEW")
    CP50.Radio_Off()
    dict = FSV.get_CSE_result()
    print(f"Conducted spurious emission test {dict['I']}")
    print(f"Marker frequency:{dict['F1']}MHz")
    print(f"Marker power:{dict['P1']}dBm")
    Result_sheet.write(row = row, column = 1, value = freq)
    Result_sheet.write(row = row, column = 2, value = sub_range)
    Result_sheet.write(row = row, column = 3, value = dict['F1'])
    Result_sheet.write(row = row, column = 4, value = dict['P1'])
    Result_sheet.write(row = row, column = 5, value = dict['F2'])
    Result_sheet.write(row = row, column = 6, value = dict['P2'])
    Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
    Result_sheet.write(row = row, column = 7, value = Timestamp)
    Result_sheet.write(row = row, column = 8, value = -30)
    Result_sheet.save("Test_Result.xlsx")


def Tx_Transient_performance(freq):

    SML.TranP_Setup(freq)
    SML.query('*OPC?')
    FSV.TranP_Setup(freq)
    FSV.query('*OPC?')
    time.sleep(2)
    FSV.write("TRIG:SOUR RFP") # set to RF power trigger
    time.sleep(2)
    FSV.write("INIT:CONT OFF")
    time.sleep(2)
    FSV.write("INIT:CONM") # set to wait for trigger
    time.sleep(2)
    CP50.Set_Freq(freq)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(1)
    CP50.Radio_Off()
    FSV.screenshot('TranP_Tx_on_frequency_'+str(freq)+'MHz')
    FSV.write("CALC:FEED 'XTIM:RFPower'") # switch to RF power screen
    time.sleep(1)
    FSV.screenshot('TranP_Tx_on_power_'+str(freq)+'MHz')
    time.sleep(3)
    FSV.write("CALC:FEED 'XTIM:FM'")
    FSV.write("INIT:CONT ON")
    FSV.write("TRIG:HOLD -0.022")# this value could be problematic
    time.sleep(3)
    CP50.Radio_On()
    time.sleep(0.1)
    CP50.Radio_Off()#Tx off transient status could be problematic
    time.sleep(2)
    FSV.screenshot('TranP_Tx_off_frequency_'+str(freq)+'MHz')
    FSV.write("CALC:FEED 'XTIM:RFPower'")
    time.sleep(1)
    FSV.screenshot('TranP_Tx_off_power_'+str(freq)+'MHz')
    FSV.write("CALC:FEED 'XTIM:FM'") # set FSV back to FM screen
    FSV.write("TRIG:SOUR IMM") # set FSV back to freerun
    SML.write(":OUTP OFF")
    print("Transient performance test completed.")

def Rx_Adjacent_channel_selectivity(freq, delta):# delta set frequency offset
    # Result_sheet.get_sheet("ACS_RAW")
    SML.Wanted_Signal(freq=freq)
    SML.query('*OPC?')
    SMB.Unwanted_Signal(freq=Decimal(freq)+Decimal(delta), unwanted_level=83.5)
    print(f"Inference frequency has been set to {Decimal(freq)+Decimal(delta)}MHz")
    SMB.query('*OPC?')
    CP50.Set_Freq(freq=Decimal(freq))
    CMS.Set_Timeout(ms=10000)
    # read initial SINAD
    SINAD_data_str = CMS.query("SINAD:R?")
    # strip number from string returned from CMS
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(SINAD_data_num)
    Level_RF = SMB.Lev_RF()
    # below code block it to test Adjacent channel selectivity
    for i in range(0,100):
        if float(SINAD_data_num) > 14.0:
            # Result_sheet.write(row = i+2, column = 1, value = Level_RF)
            # Result_sheet.write(row = i+2, column = 2, value = SINAD_data_num)
            Level_RF = Level_RF + 1
            SMB.write(f":POW {Level_RF}dBuV")
            SMB.query('*OPC?')
            SINAD_data_str = CMS.query("SINAD:R?")
            SINAD_data_num = re.findall(r'\d', SINAD_data_str)[0]
            if SINAD_data_num != '0':
                SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
                print(SINAD_data_num)
            else:
                break
        else:
            break

    ACS = Level_RF - SML.Lev_RF()
    # Result_sheet.write(row = 2, column = 10, value = ACS)
    # Result_sheet.save("Test_Result.xlsx")
    return ACS

def CHSW_ACS(start_f, start_row, end_row, clear):

    Start_F = Decimal(start_f)/Decimal(1)
    Result_sheet.get_sheet("ACS")
    if clear == 1:
        Result_sheet.clear('A2', 'F19')
    else:
        pass
    for i in range(start_row, end_row): # run whatever number of nchannels with 12.5kHz channel spacing
        ACS_high = Rx_Adjacent_channel_selectivity(freq=Decimal(Start_F), delta=0.0125)

        Result_sheet.write(row = i+2, column = 1, value = Start_F)
        Result_sheet.write(row = i+2, column = 2, value = Start_F+Decimal(0.0125))
        Result_sheet.write(row = i+2, column = 3, value = ACS_high)

        ACS_low = Rx_Adjacent_channel_selectivity(freq=Start_F, delta=-0.0125)

        Result_sheet.write(row = i+2, column = 4, value = Start_F-Decimal(0.0125))
        Result_sheet.write(row = i+2, column = 5, value = ACS_low)
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 6, value = Timestamp)

        Result_sheet.save("Test_Result.xlsx")
        Start_F += Decimal(0.0125) # frequency in MHz

    SML.write(":OUTP OFF")
    SMB.write(":OUTP1 OFF")
    print("Adjacent Channel Selectivity test completed.")


def Rx_Spurious_response_immunity(freq, delta):
    # Result_sheet.get_sheet("Spur_Res")
    SML.Wanted_Signal(freq=freq)
    SML.query('*OPC?')
    SMB.Unwanted_Signal(freq=Decimal(freq)+Decimal(delta), unwanted_level=83.5)
    print(f"Inference frequency has been set to {Decimal(freq)+Decimal(delta)}MHz")
    SMB.query('*OPC?')
    CP50.Set_Freq(freq=freq)
    CMS.Set_Timeout(ms=10000)
    SINAD_data_str = CMS.query("SINAD:R?")
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(SINAD_data_num)
    Level_RF = SMB.Lev_RF()
    # below code block is to test Spurious_Response
    for i in range(0,100):
        if float(SINAD_data_num) > 14.0:
            # Result_sheet.write(row = i+2, column = 1, value = Level_RF)
            # Result_sheet.write(row = i+2, column = 2, value = SINAD_data_num)
            Level_RF = Level_RF + 1
            SMB.write(f":POW {Level_RF}dBuV")
            SMB.query('*OPC?')
            SINAD_data_str = CMS.query("SINAD:R?")
            SINAD_data_num = re.findall(r'\d', SINAD_data_str)[0]# handle return value of 0
            if SINAD_data_num != '0':
                SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
                print(SINAD_data_num)
            else:
                break
        else:
            break

    Spur_Res = Level_RF - SML.Lev_RF()
    # Result_sheet.write(row = 2, column = 3, value = Spur_Res)
    # Result_sheet.save("Test_Result.xlsx")
    return Spur_Res


def CHSW_SR(start_f, start_row, end_row, clear):

    Start_F = Decimal(start_f)/Decimal(1) # start from 479.9875MHz
    Result_sheet.get_sheet("Spur_Res")
    if clear == 1:
        Result_sheet.clear('A2', 'M19')
    else:
        pass
    for i in range(start_row, end_row): # run first 10 channels with 12.5kHz channel spacing
        Spur_Res = Rx_Spurious_response_immunity(freq=Start_F, delta=Decimal(-2*38.85))
        Result_sheet.write(row = i+2, column = 1, value = Start_F)
        Result_sheet.write(row = i+2, column = 2, value = Start_F+Decimal(-2*38.85))
        Result_sheet.write(row = i+2, column = 3, value = Spur_Res)
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 4, value = Timestamp)
        Result_sheet.save("Test_Result.xlsx")
        Start_F += Decimal(0.0125) # frequency in MHz

    SML.write(":OUTP OFF")
    SMB.write(":OUTP1 OFF")
    print("Spurious Response rejection test completed.")


def Rx_Intermodulation_Response(freq, delta1, delta2):# delta set frequency offset
    # Result_sheet.get_sheet("Intermodulation_Response")
    SML.Wanted_Signal(freq=freq)
    SML.query('*OPC?')
    SMB.Unwanted_Signal(freq=Decimal(freq)+Decimal(delta1), unwanted_level=82.5)
    SMC.Unwanted_Signal(freq=Decimal(freq)+Decimal(delta2), unwanted_level=82.5)
    SMC.write(f":FM:STAT OFF")# turn off modulation for SMC
    print(f"Inference frequency1 has been set to {Decimal(freq)+Decimal(delta1)}MHz")
    print(f"Inference frequency2 has been set to {Decimal(freq)+Decimal(delta2)}MHz")
    SMB.query('*OPC?')
    SMC.query('*OPC?')
    CP50.Set_Freq(freq=freq)
    CMS.Set_Timeout(ms=10000)
    SINAD_data_str = CMS.query("SINAD:R?")
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(SINAD_data_num)
    Level_RF = SMB.Lev_RF()
    # below code block it to test Intermodulation Response
    for i in range(0,100):
        if float(SINAD_data_num) > 14.0:
            # Result_sheet.write(row = i+2, column = 1, value = Level_RF)
            # Result_sheet.write(row = i+2, column = 2, value = SINAD_data_num)
            Level_RF = Level_RF + 1
            SMB.write(f":POW {Level_RF}dBuV")
            SMC.write(f":POW {Level_RF}dBuV")
            SMB.query('*OPC?')
            SMC.query('*OPC?')
            SINAD_data_str = CMS.query("SINAD:R?")
            SINAD_data_num = re.findall(r'\d', SINAD_data_str)[0]
            if SINAD_data_num != '0':
                SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
                print(SINAD_data_num)
            else:
                break
        else:
            break

    Inter_Res = Level_RF - SML.Lev_RF()
    # Result_sheet.write(row = 2, column = 10, value = Inter_Res)
    # Result_sheet.save("Test_Result.xlsx")
    return Inter_Res


def CHSW_Intermodulation(start_f, start_row, end_row, clear):

    Start_F = Decimal(start_f)/Decimal(1)
    Result_sheet.get_sheet("Intermodulation_Response")
    if clear == 1:
        Result_sheet.clear('A2', 'M19')
    else:
        pass
    for i in range(start_row, end_row):
        Inter_Res_low = Rx_Intermodulation_Response(freq=Start_F, delta1=0.05, delta2=0.025)
        Result_sheet.write(row = i+2, column = 1, value = Start_F)
        Result_sheet.write(row = i+2, column = 2, value = Start_F+Decimal(0.05))
        Result_sheet.write(row = i+2, column = 3, value = Start_F+Decimal(0.025))
        Result_sheet.write(row = i+2, column = 4, value = Inter_Res_low)

        Inter_Res_high = Rx_Intermodulation_Response(freq=Start_F, delta1=0.1, delta2=0.05)
        Result_sheet.write(row = i+2, column = 5, value = Start_F+Decimal(0.1))
        Result_sheet.write(row = i+2, column = 6, value = Start_F+Decimal(0.05))
        Result_sheet.write(row = i+2, column = 7, value = Inter_Res_high)
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 8, value = Timestamp)

        Result_sheet.save("Test_Result.xlsx")
        Start_F += Decimal(0.0125) # frequency in MHz

    SML.write(":OUTP1 OFF")
    SMB.write(":OUTP1 OFF")
    SMC.write(":OUTP1 OFF")
    print("Intermodulation response rejection test completed.")

def Rx_Blocking(freq, delta):# delta set unwanted signal frequency offset
    # Result_sheet.get_sheet("Blocking")
    SML.Wanted_Signal(freq=freq)
    SML.query('*OPC?')
    SMB.Unwanted_Signal(freq=Decimal(freq)+Decimal(delta), unwanted_level=112.5)
    SMB.write(f":FM:STAT OFF")# turn off modulation for SMB
    print(f"Inference frequency has been set to {Decimal(freq)+Decimal(delta)}MHz")
    SMB.query('*OPC?')
    CP50.Set_Freq(freq=freq)
    CMS.Set_Timeout(ms=10000)
    # read initial SINAD
    SINAD_data_str = CMS.query("SINAD:R?")
    # strip number from string returned from CMS
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(SINAD_data_num)
    Level_RF = SMB.Lev_RF()

    for i in range(0,100):
        if float(SINAD_data_num) > 14.0:
            # Result_sheet.write(row = i+2, column = 1, value = Level_RF)
            # Result_sheet.write(row = i+2, column = 2, value = SINAD_data_num)
            Level_RF = Level_RF + 1
            SMB.write(f":POW {Level_RF}dBuV")
            SMB.query('*OPC?')
            SINAD_data_str = CMS.query("SINAD:R?")
            SINAD_data_num = re.findall(r'\d', SINAD_data_str)[0]
            if SINAD_data_num != '0':
                SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
                print(SINAD_data_num)
            else:
                break
        else:
            break

    BLK = Level_RF - SML.Lev_RF()
    # Result_sheet.write(row = 2, column = 10, value = BLK)
    # Result_sheet.save("Test_Result.xlsx")
    return BLK

def CHSW_BLK(start_f, start_row, end_row, clear):

    Start_F = Decimal(start_f)/Decimal(1)
    Result_sheet.get_sheet("Blocking")
    if clear == 1:
        Result_sheet.clear('A2', 'M19')
    else:
        pass
    for i in range(start_row, end_row):
        BLK_high = Rx_Blocking(freq=Start_F, delta=1)
        Result_sheet.write(row = i+2, column = 1, value = Start_F)
        Result_sheet.write(row = i+2, column = 2, value = Start_F+Decimal(1))
        Result_sheet.write(row = i+2, column = 3, value = BLK_high)
        BLK_low = Rx_Blocking(freq=Start_F, delta=-1)
        Result_sheet.write(row = i+2, column = 4, value = Start_F-Decimal(1))
        Result_sheet.write(row = i+2, column = 5, value = BLK_low)
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 6, value = Timestamp)
        Result_sheet.save("Test_Result.xlsx")
        Start_F += Decimal(0.0125) # frequency in MHz

    SML.write(":OUTP1 OFF")
    SMB.write(":OUTP1 OFF")
    print("Blocking test completed.")

def Rx_CCR(freq, delta):# delta set unwanted signal frequency offset
    # Result_sheet.get_sheet("CCR")
    SML.Wanted_Signal(freq=freq)
    SML.query('*OPC?')
    SMB.Unwanted_Signal(freq=Decimal(freq)+Decimal(delta), unwanted_level=23.5)
    print(f"Inference frequency has been set to {Decimal(freq)+Decimal(delta)}MHz")
    SMB.query('*OPC?')
    CP50.Set_Freq(freq=freq)
    CMS.Set_Timeout(ms=10000)
    # read initial SINAD
    SINAD_data_str = CMS.query("SINAD:R?")
    # strip number from string returned from CMS
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(SINAD_data_num)
    Level_RF = SMB.Lev_RF()

    for i in range(0,100):
        if float(SINAD_data_num) > 14.0:
            # Result_sheet.write(row = i+2, column = 1, value = Level_RF)
            # Result_sheet.write(row = i+2, column = 2, value = SINAD_data_num)
            Level_RF = Level_RF + 1
            SMB.write(f":POW {Level_RF}dBuV")
            SMB.query('*OPC?')
            SINAD_data_str = CMS.query("SINAD:R?")
            SINAD_data_num = re.findall(r'\d', SINAD_data_str)[0]
            if SINAD_data_num != '0':
                SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
                print(SINAD_data_num)
            else:
                break
        else:
            break

    CCR = Level_RF - SML.Lev_RF()
    # Result_sheet.write(row = 2, column = 10, value = CCR)
    # Result_sheet.save("Test_Result.xlsx")
    return CCR

def CHSW_CCR(start_f, start_row, end_row, clear):

    Start_F = Decimal(start_f)/Decimal(1)
    Result_sheet.get_sheet("CCR")
    if clear == 1:
        Result_sheet.clear('A2', 'M19')
    else:
        pass
    for i in range(start_row, end_row):
        Result_sheet.write(row = i+2, column = 1, value = Start_F)
        CCR = Rx_CCR(freq=Start_F, delta=0)
        Result_sheet.write(row = i+2, column = 2, value = Start_F+Decimal(0))
        Result_sheet.write(row = i+2, column = 3, value = CCR)
        CCR_high = Rx_CCR(freq=Start_F, delta=0.0015)
        Result_sheet.write(row = i+2, column = 4, value = Start_F+Decimal(0.0015))
        Result_sheet.write(row = i+2, column = 5, value = CCR_high)
        CCR_low = Rx_CCR(freq=Start_F, delta=-0.0015)
        Result_sheet.write(row = i+2, column = 6, value = Start_F-Decimal(0.0015))
        Result_sheet.write(row = i+2, column = 7, value = CCR_low)
        CCR_high = Rx_CCR(freq=Start_F, delta=0.003)
        Result_sheet.write(row = i+2, column = 8, value = Start_F+Decimal(0.003))
        Result_sheet.write(row = i+2, column = 9, value = CCR_high)
        CCR_low = Rx_CCR(freq=Start_F, delta=-0.003)
        Result_sheet.write(row = i+2, column = 10, value = Start_F-Decimal(0.003))
        Result_sheet.write(row = i+2, column = 11, value = CCR_low)
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 12, value = Timestamp)
        Result_sheet.save("Test_Result.xlsx")
        Start_F += Decimal(0.0125) # frequency in MHz

    SML.write(":OUTP1 OFF")
    SMB.write(":OUTP1 OFF")
    print("Co-channel response test completed.")

def Rx_Maxmium_usable_sensitivity(freq):
    # Result_sheet.get_sheet("MUS")
    SML.Wanted_Signal(freq=freq)
    SML.query('*OPC?')
    CP50.Set_Freq(freq=freq)
    CMS.Set_Timeout(ms=10000)
    # read initial SINAD
    SINAD_data_str = CMS.query("SINAD:R?")
    # strip number from string returned from CMS
    SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
    print(SINAD_data_num)
    Level_RF = SML.Lev_RF()

    for i in range(0,100):
        if float(SINAD_data_num) > 20.0:
            # Result_sheet.write(row = i+2, column = 1, value = Level_RF)
            # Result_sheet.write(row = i+2, column = 2, value = SINAD_data_num)
            Level_RF = Level_RF - 1
            SML.write(f":POW {Level_RF}dBuV")
            SML.query('*OPC?')
            SINAD_data_str = CMS.query("SINAD:R?")
            SINAD_data_num = re.findall(r'\d', SINAD_data_str)[0]
            if SINAD_data_num != '0':
                SINAD_data_num = re.findall(r'\d+\.\d+', SINAD_data_str)[0]
                print(SINAD_data_num)
            else:
                break
        else:
            break

    MUS = Level_RF-SML.Lev_RF()
    # Result_sheet.write(row = 2, column = 10, value = MUS)
    # Result_sheet.save("Test_Result.xlsx")
    return MUS

def CHSW_MUS(start_f, start_row, end_row, clear):

    Start_F = Decimal(start_f)/Decimal(1)
    Result_sheet.get_sheet("MUS")
    if clear == 1:
        Result_sheet.clear('A2', 'M19')
    else:
        pass
    for i in range(start_row, end_row):
        MUS = Rx_Maxmium_usable_sensitivity(freq=Start_F)
        Result_sheet.write(row = i+2, column = 1, value = Start_F)
        Result_sheet.write(row = i+2, column = 2, value = MUS)
        Timestamp ='{:%d-%b-%Y %H:%M:%S}'.format(datetime.datetime.now())
        Result_sheet.write(row = i+2, column = 3, value = Timestamp)
        Result_sheet.save("Test_Result.xlsx")
        Start_F += Decimal(0.0125) # frequency in MHz

    SML.write(":OUTP1 OFF")
    SMB.write(":OUTP1 OFF")
    print("Maximum usable sensitivity(conducted) test completed.")


try:
    SC = SoundCard()
except BaseException:
    print("Specified Soundcard does not exist.")
    pass

try:
    SW1 = SwitchBox('11108140003','A')
except BaseException:
    print("Specified SwitchBox does not exist.")
    pass

try:
    FSV = SpecAn('TCPIP0::10.0.22.42::inst0::INSTR')
except BaseException:
    print("FSV is not on.")
    pass

try:
    CP50 = Radio('com6')
except BaseException:
    print("Specified com port does not exsit.")
    pass

try:
    SML = SigGen('TCPIP0::10.0.22.94::inst0::INSTR')
except BaseException:
    print("SML is not on.")
    pass

try:
    SMB = SigGen('TCPIP0::10.0.22.100::inst0::INSTR')
except BaseException:
    print("SMB is not on.")
    pass

try:
    SMC = SigGen('TCPIP0::10.0.22.102::inst0::INSTR')
except BaseException:
    print("SMC is not on.")
    pass

try:
    CMS = SigGen('GPIB0::24::INSTR')
except BaseException:
    print("CMS is not on.")
    pass

try:
    Result_sheet = Excel("Test_Result.xlsx")
    #CP50_Result = Excel("CP50_Result.xlsx")
except BaseException:
    print("Specified Excel file does not exsit.")
    pass



# SC.live_plots()
# Tx_Adjacent_channel_power(459.075)



# Tx_Max_deviation_CB(476.9)
# Tx_Frequency_error_Carrier_power(459.075)
# Tx_Conducted_spurious_emissions()
# Tx_Occupied_bandwidth(472.0125)
# Tx_Out_of_band_modulation_response(476.9)
# Tx_Transient_performance(476.9)

# FSV.DeMod_Setup(472.0125)
# SML.Tx_Setup()
# CP50.Set_Freq(476.9)
# CP50.Set_Pow("high")
# CP50.Radio_On()
# time.sleep(120)
# CP50.Radio_Off()

# FSV.screenshot('OCW02')

#FSV.screenshot('EIRP02')

# Tx_Occupied_bandwidth(472.025, 'bottom')
# Tx_Occupied_bandwidth(472.1, 'top')

# CHSW_MUS(451, 0)
# CHSW_MUS(485, 4)
# CHSW_MUS(519, 8)
# CHSW_CCR(451, 0)
# CHSW_CCR(485, 4)
# CHSW_CCR(519, 8)
# CHSW_BLK(451.0125, 0)
# CHSW_BLK(485, 4)
# CHSW_BLK(519, 8)
# CHSW_ACS(459.075, start_row=0, end_row=1, clear=1)
# CHSW_ACS(485, 4)
# CHSW_ACS(519, 8)
# CHSW_SR(451, 12)
# CHSW_SR(485, 16)
# CHSW_SR(519, 20)
# CHSW_Intermodulation(451.0125, 0)
#CHSW_Intermodulation(485, 4)
#CHSW_Intermodulation(519, 8)
