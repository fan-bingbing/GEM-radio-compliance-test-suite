import sys
import math
import visa
from openpyxl import load_workbook
import time
import serial

# below codes are for importing files from different directory
# a __init__.py file has to be exist in both "import from" and "import to" folder
sys.path.insert(0, r'C:\Users\afan\documents\commercial-radio-test-suite\midlevel')
sys.path.insert(0, r'C:\Users\afan\documents\commercial-radio-test-suite\bottomlevel')
# above codes are for importing files from different directory
# a __init__.py file has to be exist in both "import from" and "import to" folder

# import config # containing global variables
# import Rx_Write_to_excel as RxW
# import FSV_screenshot as Scr
# import RadioControl as Rcon
# import FreqError_Power as FEP
# import Max_Deviation as MAD
# import OOB_ModRes as OOB
# import ACP
# import Cond_Spur as Cons
# import ACS
# import Blocking as BLK
# import Spur_Res as SPR
#
#

    # Test_frequency = input("Input test frequency in Mhz or press CTRL+C to quit > ")
    # Rcon.RadioControl('com7', 'ON')
    # dict = FEP.FreqError_Power(Test_frequency)
    # print(f"Frequency error and Carrier power test {dict['Indication']}")
    # print(f"Frequency error:{dict['Frequency_error']}Hz")
    # print(f"Carrier power:{dict['Carrier_power']}dBm")
    # Rcon.RadioControl('com7', 'OFF')
    # file_name = input("To save the screenshot, input the filename (***.bmp) or press CTRL+C to quit > ")
    # Scr.Screenshot(file_name)


class SpecAn(object):
    def __init__(self):

        self.rm = visa.ResourceManager()
        self.SP = self.rm.open_resource('TCPIP0::192.168.10.9::hislip0::INSTR') # Spec An
        self.book = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file

    def FEP_Setup(self, freq):

        self.sheet = self.book["Freq_Error_Power"] # load excel sheet

        # below codes are for setting test frequency in Test_Setup.xlsx according to user's input
        self.sheet.cell(row = 1, column = 2, value = freq) # write test frequency in this self.sheet
        self.book.save("Test_Setup.xlsx") # save existing .xlsx file
        # above codes are for setting test frequency in Test_Setup.xlsx according to user's input

        self.Centre_frequency = self.sheet["B1"].value #
        self.Span_frequency = self.sheet["B2"].value #
        self.RBW = self.sheet["B3"].value #
        self.VBW = self.sheet["B4"].value #
        self.RF_level = self.sheet["B5"].value #
        self.Attenuation = self.sheet["B6"].value #
        self.RefLev_offset = self.sheet["B7"].value#
        self.Trace_peak = self.sheet["B8"].value #
        self.Transducer1 = self.sheet["B9"].value # cable
        self.Trans1_ON = self.sheet["B10"].value
        self.Transducer2 = self.sheet["B11"].value # 30dB attenuator
        self.Trans2_ON = self.sheet["B12"].value
        self.Transducer3 = self.sheet["B13"].value # High pass filter
        self.Trans3_ON = self.sheet["B14"].value
        self.Limit_line_1 = self.sheet["B15"].value # ASNZS4365:2011
        self.Limit_line_1_ON = self.sheet["B16"].value # ASNZS4365:2011
        self.Limit_line_2 = self.sheet["B17"].value # ASNZS4295:2015
        self.Limit_line_2_ON = self.sheet["B18"].value # ASNZS4295:2015
        self.Sweep_points= self.sheet["B19"].value # get sweep points

        self.SP.write(f"*RST")
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write(f"FREQ:CENT {self.Centre_frequency}MHz")
        self.SP.write(f"FREQ:SPAN {self.Span_frequency}Hz")
        self.SP.write(f"BAND {self.RBW}Hz")
        self.SP.write(f"BAND:VID {self.VBW}Hz")
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {self.RefLev_offset}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {self.RF_level}")
        self.SP.write(f"INP:ATT {self.Attenuation}")
        self.SP.write(f"{self.RefLev_offset}")
        self.SP.write(f"{self.Trace_peak}")
        self.SP.write(f"CORR:TRAN:SEL '{self.Transducer1}'")
        self.SP.write(f"CORR:TRAN {self.Trans1_ON} ")
        self.SP.write(f"CORR:TRAN:SEL '{self.Transducer2}'")
        self.SP.write(f"CORR:TRAN {self.Trans2_ON}")
        self.SP.write(f"CORR:TRAN:SEL '{self.Transducer3}'")
        self.SP.write(f"CORR:TRAN {self.Trans3_ON}")
        self.SP.write(f"CALC:LIM:NAME '{self.Limit_line_1}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {self.Limit_line_1_ON}")
        self.SP.write(f"CALC:LIM:NAME '{self.Limit_line_2}'")
        self.SP.write(f"CALC:LIM:UPP:STAT {self.Limit_line_2_ON}")
        self.SP.write(f"SWE:POIN {self.Sweep_points}")


    def MAD_Setup(self, freq):
        self.sheet = self.book["Max_Deviation"]
        # below codes are for setting test frequency in Test_Setup.xlsx according to user's input
        self.sheet.cell(row = 1, column = 2, value = freq) # write test frequency in this self.sheet
        self.book.save("Test_Setup.xlsx") # save existing .xlsx file
        # above codes are for setting test frequency in Test_Setup.xlsx according to user's input

        # following code is to initialize FSV
        self.Centre_frequency = self.sheet["B1"].value #
        self.Dev_PerDivision = self.sheet["B2"].value #
        self.Demod_BW = self.sheet["B3"].value #
        self.AF_Couple = self.sheet["B4"].value #
        self.RF_level = self.sheet["B5"].value #
        self.Attenuation = self.sheet["B6"].value # get attenuation
        self.RefLev_offset = self.sheet["B7"].value# get RFlevel offset
        self.Trace_Peak = self.sheet["B8"].value # get trace to Pos peak
        self.Demod_MT = self.sheet["B9"].value #
        self.Cont_sweep = self.sheet["B10"].value #
        self.SP.write(f"*RST")
        self.SP.write("SYST:DISP:UPD ON")
        self.SP.write("ADEM ON")
        self.SP.write(f"FREQ:CENT {self.Centre_frequency}MHz")
        self.SP.write(f"DISP:TRAC:Y:PDIV {self.Dev_PerDivision}kHz")
        self.SP.write(f"BAND:DEM {self.Demod_BW}kHz")
        self.SP.write(f"ADEM:AF:COUP {self.AF_Couple}")#Set AF coupling to AC, the frequency offset is automatically corrected.
        # i.e. the trace is always symmetric with respect to the zero line
        self.SP.write(f"DISP:TRAC:Y:RLEV:OFFS {self.RefLev_offset}")
        self.SP.write(f"DISP:TRAC:Y:RLEV {self.RF_level}")
        self.SP.write(f"INP:ATT {self.Attenuation}")
        self.SP.write(f"{self.Trace_Peak}")
        self.SP.write(f"ADEM:MTIM {self.Demod_MT}ms")
        self.SP.write(f"INIT:CONT {self.Cont_sweep}")
        # above code is to initialize FSV



    def write(self, str):
        self.SP.write(str)

    def query(self, str):
        return self.SP.query(str)


    def get_FEP_result(self):
        self.SP.write("CALC:MARK1:MAX")
        self.Frequency = self.SP.query("CALC:MARK1:X?")
        self.Level = float(self.SP.query("CALC:MARK1:Y?"))
        self.Frequency_error = float(self.Frequency) - float(self.Centre_frequency)*1e6
        self.indication = (self.SP.query("*OPC?")).replace("1","Completed.")

        return {'F':self.Frequency_error, 'P':self.Level, 'I':self.indication}


    def close(self):
        self.SP.close()

class SigGen(object):

    def __init__(self):
        self.rm = visa.ResourceManager()
        self.SML = self.rm.open_resource('ASRL4::INSTR')
        self.book = load_workbook(filename = "Test_Setup.xlsx") # load an existing .xlsx file
        self.sheet = self.book["Max_Deviation"]
        # following code is to initialize SML
        self.Frequency_AF = self.sheet["G1"].value #
        self.Level_AF = self.sheet["G2"].value #
        self.AF_output_on = self.sheet["G3"].value #


    def MAD_Setup(self):


        self.SML.write(f"*RST")
        self.SML.write("SYST:DISP:UPD ON")
        self.SML.write(f":FM:INT:FREQ {self.Frequency_AF}kHz")
        self.SML.write(f":OUTP2:VOLT {self.Level_AF}mV")
        self.SML.write(f":OUTP2 {self.AF_output_on}")

    def Lev_AF(self):
        return self.Level_AF # useful for Max_deviation test

    def write(self, str):
        self.SML.write(str)

    def query(self, str):
        return self.SML.query(str)

    def close(self):
        self.SML.close()


class Radio(object):
    def __init__(self):
        self.ESC_CHAR = 0x7D # Escape char
        self.STX_CHAR =  0x7E # Start of packet
        self.ETX_CHAR = 0x7F # End of packet
        self.CHECKSUM_XOR_MASK = 0xFF

        self.payload0 = bytearray(b'\xB4\x88\x2a\x80')# set carrie freq 136MHz basd on 12.5kHz calculation
        self.payload1 = bytearray(b'\xB4\x93\x03')# set 25W power
        self.payload2 = bytearray(b'\xA6\x01') # PTT ON
        self.payload3 = bytearray(b'\xB4\x91\x03\xE8')# 1kHz audio on
        self.payload4 = bytearray(b'\xB4\x91\x0B\xB8')# 3kHz audio on
        self.payload5 = bytearray(b'\xB4\x91\x00\x00')# audio off
        self.payload6 = bytearray(b'\xB4\x82\x10')# selcall tone on
        self.payload7 = bytearray(b'\xB4\x82\x0f')# selcall tone off
        self.payload8 = bytearray(b'\xA6\x00') # PTT off

        self.port = "com7"
        self.baudrate = 115200
        self.timeout = None
        self.ser = serial.Serial(port=self.port, baudrate=self.baudrate, timeout=self.timeout)  # open serial port


    def Packet_Gen(self, payload):
        # Header
        tx_array = bytearray((self.STX_CHAR, 2, len(payload)))

        # Escape the payload, append it to the output buffer, make the checksum
        chksum = 0
        for i, b in enumerate(payload):
            # The first word of the payload is always the command ID.
            # The MSB of the command ID is always 1. The radio just
            # clears it in the ACK, without doing anything else with it.
            if i == 0:
                bb = (0x80 | b) & 0xFF
            else:
                bb = b & 0xFF
            chksum ^= bb
            if bb in (self.ESC_CHAR, self.STX_CHAR, self.ETX_CHAR):
                tx_array.append(self.ESC_CHAR)
                tx_array.append(bb ^ 0x20)
            else:
                tx_array.append(bb)

        # Trailer
        tx_array.append(chksum ^ self.CHECKSUM_XOR_MASK)
        tx_array.append(self.ETX_CHAR)

        return tx_array

    def Set_Freq(self, freq):
        a = ((hex(int((freq*1e6)/(12.5*1e3)))[2]+hex(int((freq*1e6)/(12.5*1e3)))[3]))# calculate first HEX of input frequency
        b = ((hex(int((freq*1e6)/(12.5*1e3)))[4]+hex(int((freq*1e6)/(12.5*1e3)))[5]))# calculate second HEX of input frequency
        self.payload0[2] = int(a,16) # assign first DEC (transferred from HEX) to the third number of paylaod0
        self.payload0[3] = int(b,16) # assign second DEC (transferred from HEX) to the fourth number of paylaod0
        self.ser.write(self.Packet_Gen(self.payload0))
        print(f"radio frequency has been set to {freq} MHz.")

    def Set_Pow(self, pow):
        if pow == "low":
            self.payload1[2] = 0
        elif pow == "high":
            self.payload1[2] = 3
        else:
            print("Set_Pow() only accept 'low' or 'high'")
        print(f"radio power has been set to {pow}.")

        self.ser.write(self.Packet_Gen(self.payload1))

    def Radio_On(self):
        self.ser.write(self.Packet_Gen(self.payload2))
        print("Radio's on")

    def Radio_Off(self):
        self.ser.write(self.Packet_Gen(self.payload8))
        print("Radio's off")



    # def Radio_Control(self):
    #     # time.sleep(0.5)
    #     self.ser.write(self.Packet_Gen(payload))
    #     print("one command sent.") # set a indicator showing one command has been sent to radio

    def Radio_close(self):
        self.ser.close()
        print("Serial session is closed.")


def Tx_Frequency_error_Carrier_power():

    RFile_write = load_workbook(filename = "Test_Result.xlsx") # load Test_Result.xlsx
    RSheet = RFile_write["Ferror_Pow"] # load "Ferror_Pow" sheet in .xlsx
    FSV.FEP_Setup(470)
    FSV.query('*OPC?')


    CP50.Set_Freq(470)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(2)
    FSV.write("DISP:TRAC:MODE MAXH")
    time.sleep(3)
    FSV.write("DISP:TRAC:MODE VIEW")
    CP50.Radio_Off()
    dict = FSV.get_FEP_result()
    print(f"Frequency error and Carrier power test {dict['I']}")
    print(f"Frequency error:{dict['F']}Hz")
    print(f"Carrier power:{dict['P']}dBm")

    RSheet.cell(row = 2, column = 1, value = dict['F'])
    RSheet.cell(row = 2, column = 2, value = dict['P'])
    RFile_write.save("Test_Result.xlsx")

    FSV.close()
    CP50.Radio_close()


def Tx_Max_deviation():
    AF_list = [100, 300, 500, 700, 900, 1000, 1300, 1500, 1700, 1900, 2000, 2300, 2550] # audio frequency list
    Reading_list = [] # empty list for deviation result storage
    RFile_write = load_workbook(filename = "Test_Result.xlsx") # load Test_Result.xlsx
    RSheet = RFile_write["Max_Dev"] # load "Ferror_Pow" sheet in .xlsx
    SML.MAD_Setup()
    SML.query('*OPC?')
    FSV.MAD_Setup(470)
    FSV.query('*OPC?')

    CP50.Set_Freq(470)
    CP50.Set_Pow("high")
    CP50.Radio_On()
    time.sleep(2)

    FSV.write(f"INIT:CONT OFF")

    Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0 #get the initial deviation value


    Level_AF = SML.Lev_AF()# initial Level_AF

    if Dev_Reading < 1.5:
        while Dev_Reading < 1.47:
            Level_AF = Level_AF+1
            SML.write(f":OUTP2:VOLT {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0
    else:
        while Dev_Reading > 1.53:
            Level_AF = Level_AF-1
            SML.write(f":OUTP2:VOLT {Level_AF}mV")
            FSV.write(f"INIT:CONT ON")
            time.sleep(1)
            FSV.write(f"INIT:CONT OFF")
            time.sleep(1)
            Dev_Reading = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000.0
    # above code is to find Audio output level satisfying standard condtion (around 1.5kHz deviation)

    FSV.write(f"INIT:CONT ON")
    SML.query('*OPC?')


    #following code is to vary audio frequency to complete the test
    Level_AF = 100*Level_AF # bring audio level up 20dB in one step according to standard
    SML.write(f":OUTP2:VOLT {Level_AF}mV")

    for i in range(0,13):
        print(f"At Audio frequency:{AF_list[i]}")
        SML.write(f":FM:INT:FREQ {AF_list[i]}Hz")
        SML.query('*OPC?')
        FSV.write(f"INIT:CONT OFF")
        time.sleep(1)
        Dev = float(FSV.query("CALC:MARK:FUNC:ADEM:FM? MIDD"))/1000
        Reading_list.append(Dev)
        FSV.query("*OPC?")
        FSV.write(f"INIT:CONT ON")
        print(f"Deviation is {Reading_list[i]}kHz")
        RSheet.cell(row = i+2, column = 1, value = AF_list[i])
        RSheet.cell(row = i+2, column = 2, value = Dev)


    RFile_write.save("Test_Result.xlsx") # save existing .xlsx file
    CP50.Radio_Off()
    SML.write(":OUTP2 OFF")# turn off audio output at the end of the test
    indication = (FSV.query("*OPC?")).replace("1","Completed")
    print(f"Maximum Deviation test {indication}.")

    FSV.close()
    SML.close()
    CP50.Radio_close()



FSV = SpecAn()
CP50 = Radio()
SML = SigGen()



    # FSV.write("CALC:MARK1:MAX")
    # print(FSV.query("CALC:MARK1:X?"))








    #print(FSV.query("*OPC?"))

    #FSV.close()
     # replace return character "1" to "completed"
    # RSheet.cell(row = 2, column = 1, value = dict['F'])
    # RSheet.cell(row = 2, column = 2, value = dict['P'])
    # RFile_write.save("Test_Result.xlsx") # save existing .xlsx file
    #
    # print(f"Frequency error and Carrier power test {dict['I']}")
    # print(f"Frequency error:{dict['F']}Hz")
    # print(f"Carrier power:{dict['P']}dBm")







#
#
# def Tx_Max_deviation():
#     Test_frequency = input("Input test frequency in Mhz or press CTRL+C to quit > ")
#     Rcon.RadioControl('com7', 'ON')
#     list = MAD.Max_Deviation(Test_frequency)
#     print(f"Max deviation test {list[13]}")
#     print(f"Result stored in:{list}")
#     Rcon.RadioControl('com7', 'OFF')
#
#
# def Tx_Adjacent_channel_power():
#     Test_frequency = input("Input test frequency in Mhz or press CTRL+C to quit > ")
#     Rcon.RadioControl('com7', 'ON')
#     print(ACP.Adjacent_channel_power(Test_frequency))
#     Rcon.RadioControl('com7', 'OFF')
#     file_name = input("To save the screenshot, input the filename (***.bmp) or press CTRL+C to quit > ")
#     Scr.Screenshot(file_name)
#
#
# def Tx_Transient_performance():
#     print("blank")
#
#
# def Tx_Out_of_band_modulation_response():
#     Test_frequency = input("Input test frequency in Mhz or press CTRL+C to quit > ")
#     Rcon.RadioControl('com7', 'ON')
#     Abs_deviation = OOB.OOB_ModRes(Test_frequency)
#     Rel_deviation = []
#     for i in range(0,8):
#         Rel_deviation.append(20*math.log10(Abs_deviation[i]/Abs_deviation[8]))
#     print(f"Out of Band Modulation Response test {Abs_deviation[9]}")
#     print(f"Absolution deviation:{Abs_deviation}")
#     print(f"Relative deviation:{Rel_deviation}")
#     Rcon.RadioControl('com7', 'OFF')
#
#
#
# def Tx_Conducted_spurious_emissions():
#     print("--------------------------------------------")
#     print("Conducted_spurious_emissions_test_menu")
#     print("--------------------------------------------")
#     print("1. 9k-150kHz")
#     print("2. 150kHz-30MHz")
#     print("3. 30MHz-700MHz")
#     print("4. 700MHz-1GHz")
#     print("5. 1-4GHz")
#     print("0. Exit")
#     while True:
#         choice = input("> ")
#         if choice == "1":
#             print("Check physical test setup: 30dB attenuator in line WITHOUT high pass filter ?")
#             dict = Conducted_spurious_menu("Cond_Spurious_1")
#             print(f"Mark frequency:{dict['Mark frequency']/1e3}kHz")
#             print(f"Mark level:{dict['Mark level']}dBm")
#             break
#         elif choice == "2":
#             print("Check physical test setup: 30dB attenuator in line WITHOUT high pass filter ?")
#             dict = Conducted_spurious_menu("Cond_Spurious_2")
#             print(f"Mark frequency:{dict['Mark frequency']/1e6}MHz")
#             print(f"Mark level:{dict['Mark level']}dBm")
#             break
#         elif choice == "3":
#             print("Check physical test setup: 30dB attenuator in line WTIHOUT high pass filter ?")
#             dict = Conducted_spurious_menu("Cond_Spurious_3")
#             print(f"Mark frequency:{dict['Mark frequency']/1e6}MHz")
#             print(f"Mark level:{dict['Mark level']}dBm")
#             break
#         elif choice == "4":
#             print("Check physical test setup: 30dB attenuator in line WITH high pass filter NHP700+ ?")
#             dict = Conducted_spurious_menu("Cond_Spurious_4")
#             print(f"Mark frequency:{dict['Mark frequency']/1e6}MHz")
#             print(f"Mark level:{dict['Mark level']}dBm")
#             break
#         elif choice == "5":
#             print("Check physical test setup: 30dB attenuator in line WITH high pass filter NHP700+ ?")
#             dict = Conducted_spurious_menu("Cond_Spurious_5")
#             print(f"Mark frequency:{dict['Mark frequency']/1e6}MHz")
#             print(f"Mark level:{dict['Mark level']}dBm")
#             break
#
#         elif choice == "0":
#             exit(0)
#         else:
#             print("please enter number within the range from 0 to 5.")
#
#     Tx_Conducted_spurious_emissions() # go back to conducted spurious menu after break
#
#
#
# def Conducted_spurious_menu(frequency_range):
#
#     print("1. YES, Continue")
#     print("2. NO, Go back to previous menu")
#     while True:
#         choice = input("> ")
#         if choice == "1":
#             Rcon.RadioControl('com7', 'ON')
#             dict = Cons.Conducted_spurious(frequency_range)
#             print(f"Conducted_spurious_emissions_test {dict['Indication']}")
#             #print(f"Mark frequency:{dict['Mark frequency']}Hz")
#             #print(f"Mark level:{dict['Mark level']}dBm")
#             Rcon.RadioControl('com7', 'OFF')
#             file_name = input("To save the screenshot, input the filename (***.bmp) or press CTRL+C to quit > ")
#             Scr.Screenshot(file_name)
#             return dict
#             break
#
#         elif choice == "2":
#             Tx_Conducted_spurious_emissions()
#         else:
#             print("please enter number within the range from 1 or 2.")
#
#
# def Rx_Spurious_emissions():
#     print("blank")
#
#
# def Rx_Adjacent_channel_selectivity():
#     Test_frequency = input("Input test frequency in Mhz or press CTRL+C to quit > ")
#     dict = ACS.Adjacent_channel_selectivity(Test_frequency)
#     print (dict['Indication'])
#     print (f"ACS+: {dict['ACS_high']}, ACS-: {dict['ACS_low']}")
#
#
# def Rx_Spurious_response_immunity():
#     Test_frequency = input("Input test frequency in Mhz or press CTRL+C to quit > ")
#     dict = SPR.Spurious_response_immunity(Test_frequency)
#     print (dict['Indication'])
#     print (f"Spurious response at 2*IF offset: {dict['Spur_Res']}")
#
#
#
# def Rx_Blocking_immunity():
#     Test_frequency = input("Input test frequency in Mhz or press CTRL+C to quit > ")
#     dict = BLK.Blocking_immunity(Test_frequency)
#     print (dict['Indication'])
#     print (f"BLK+: {dict['BLK_high']}, BLK-: {dict['BLK_low']}")
